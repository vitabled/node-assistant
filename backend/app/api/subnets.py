"""API «Подсети» (Обходы БС, Wave-5 PR-5). Хранилище — services/subnets_store."""
import asyncio
import csv
import io
import json
import time

import httpx
import openpyxl
from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse, Response
from openpyxl.styles import Font
from pydantic import BaseModel, Field, model_validator

from app.services import latency_lab
from app.services import asn_store
from app.services import subnets_store as store

router = APIRouter(prefix="/api/subnets")


def _not_found(exc: KeyError) -> HTTPException:
    return HTTPException(404, f"Не найдено: {exc}")


@router.get("")
async def get_all():
    return {**store.get_store(), "operators": store.OPERATORS}


# ── справочник ASN (per-account, синхронизация asnname в строках) ──
class AsnBody(BaseModel):
    asn: str
    name: str = ""
    note: str = ""


@router.get("/asns")
async def list_asns():
    return {"ok": True, "asns": asn_store.list_asns()}


@router.post("/asns")
async def upsert_asn(body: AsnBody):
    """Создать/обновить запись справочника (asn нормализуется: «12345» →
    «AS12345»). После upsert у ВСЕХ строк подсетей текущего аккаунта с
    values.asn == «AS12345» перезаписывается values.asnname = name —
    справочник авторитетнее ручной правки ячейки."""
    try:
        rec = asn_store.upsert_asn(body.asn, body.name, body.note)
    except ValueError as e:
        raise HTTPException(422, str(e))
    updated_rows = store.apply_asn_name(rec["asn"], rec["name"])
    return {"ok": True, "asn": rec, "updated_rows": updated_rows}


@router.delete("/asns/{asn}")
async def delete_asn(asn: str):
    """Удалить запись справочника. asnname в строках подсетей НЕ трогается —
    остаётся последнее название из справочника. Файл иконки записи удаляется."""
    try:
        key = asn_store.normalize_asn(asn)
    except ValueError as e:
        raise HTTPException(422, str(e))
    asn_store.delete_asn(key)
    return {"ok": True}


# ── иконки записей ASN (у ASN, а не у файлов/провайдеров) ─────
@router.post("/asns/{asn}/icon")
async def upload_asn_icon(asn: str, file: UploadFile = File(...)):
    """Загрузить иконку записи ASN (png/svg/webp ≤ 256 КБ, multipart).
    Иконка сама подтягивается к подсетям с этим ASN (GET /asns/{asn}/icon)."""
    try:
        key = asn_store.normalize_asn(asn)
    except ValueError as e:
        raise HTTPException(422, str(e))
    try:
        asn_store.save_asn_icon(key, await file.read(), file.filename or "")
    except KeyError as e:
        raise _not_found(e)
    except ValueError as e:
        raise HTTPException(400, str(e))
    return {"ok": True}


@router.get("/asns/{asn}/icon")
async def get_asn_icon(asn: str):
    """Файл иконки записи ASN (404 — записи нет или иконка не загружена)."""
    try:
        key = asn_store.normalize_asn(asn)
    except ValueError as e:
        raise HTTPException(422, str(e))
    try:
        path = asn_store.asn_icon_file(key)
    except KeyError as e:
        raise _not_found(e)
    if path is None:
        raise HTTPException(404, "Иконка не загружена")
    return FileResponse(path)


class ProviderBody(BaseModel):
    name: str = ""


@router.post("/providers", status_code=201)
async def create_provider(body: ProviderBody):
    return store.add_provider(body.name)


@router.patch("/providers/{provider_id}")
async def rename_provider(provider_id: str, body: ProviderBody):
    try:
        store.rename_provider(provider_id, body.name)
    except KeyError as e:
        raise _not_found(e)
    return {"ok": True}


@router.delete("/providers/{provider_id}", status_code=204)
async def delete_provider(provider_id: str):
    store.delete_provider(provider_id)


class ListBody(BaseModel):
    name: str = ""


@router.post("/providers/{provider_id}/lists", status_code=201)
async def create_list(provider_id: str, body: ListBody):
    try:
        return store.add_list(provider_id, body.name)
    except KeyError as e:
        raise _not_found(e)


@router.patch("/providers/{provider_id}/lists/{list_id}")
async def rename_list(provider_id: str, list_id: str, body: ListBody):
    try:
        store.rename_list(provider_id, list_id, body.name)
    except KeyError as e:
        raise _not_found(e)
    return {"ok": True}


@router.delete("/providers/{provider_id}/lists/{list_id}", status_code=204)
async def delete_list(provider_id: str, list_id: str):
    try:
        store.delete_list(provider_id, list_id)
    except KeyError as e:
        raise _not_found(e)


class RowsBody(BaseModel):
    """Строки списка: старый формат {subnets: [...]} или строки с метаданными
    {rows: [{subnet, ...поля}]}. Хотя бы один из списков непустой."""
    subnets: list[str] = []
    rows: list[dict] = []

    @model_validator(mode="after")
    def _at_least_one(self):
        if not self.subnets and not self.rows:
            raise ValueError("Передайте subnets (подсети) или rows (строки с метаданными)")
        return self


@router.post("/providers/{provider_id}/lists/{list_id}/rows", status_code=201)
async def add_rows(provider_id: str, list_id: str, body: RowsBody):
    try:
        items = body.subnets if body.subnets else body.rows
        return store.add_rows(provider_id, list_id, items)
    except KeyError as e:
        raise _not_found(e)


@router.delete("/providers/{provider_id}/lists/{list_id}/rows/{row_id}", status_code=204)
async def delete_row(provider_id: str, list_id: str, row_id: str):
    try:
        store.delete_row(provider_id, list_id, row_id)
    except KeyError as e:
        raise _not_found(e)


class CellBody(BaseModel):
    value: str = ""


@router.patch("/providers/{provider_id}/lists/{list_id}/rows/{row_id}/cell/{col_key}")
async def set_cell(provider_id: str, list_id: str, row_id: str, col_key: str, body: CellBody):
    try:
        store.set_cell(provider_id, list_id, row_id, col_key, body.value)
    except KeyError as e:
        raise _not_found(e)
    return {"ok": True}


# ── пакетное обновление ячеек ──────────────────────────────────
MAX_BATCH_UPDATES = 500


class CellUpdateBody(BaseModel):
    row_id: str
    col: str
    value: str = ""


class BatchCellsBody(BaseModel):
    updates: list[CellUpdateBody] = Field(..., min_length=1, max_length=MAX_BATCH_UPDATES)


@router.patch("/providers/{provider_id}/lists/{list_id}/rows/batch")
async def batch_update_cells(provider_id: str, list_id: str, body: BatchCellsBody):
    """Пакетное обновление ячеек. Битые апдейты (нет строки/пустой col)
    не убивают batch: считаются в skipped и перечисляются в errors."""
    _pick_list(provider_id, list_id)  # 404, если списка нет
    updated, skipped = 0, 0
    errors: list[str] = []
    for u in body.updates:
        if not (u.row_id or "").strip() or not (u.col or "").strip():
            skipped += 1
            errors.append("row_id/col не могут быть пустыми")
            continue
        try:
            store.set_cell(provider_id, list_id, u.row_id, u.col, u.value)
            updated += 1
        except KeyError as e:
            skipped += 1
            errors.append(f"row {u.row_id}: не найдена ({e})")
        except Exception as e:  # битая строка не убивает batch
            skipped += 1
            errors.append(str(e))
    return {"ok": True, "updated": updated, "skipped": skipped, "errors": errors}


# ── JSON-импорт строк с метаданными ────────────────────────────
class ImportJsonBody(BaseModel):
    rows: list[dict] = Field(..., min_length=1)


@router.post("/providers/{provider_id}/lists/{list_id}/import-json")
async def import_json_rows(provider_id: str, list_id: str, body: ImportJsonBody):
    """Импорт строк {subnet, ...метаданные}. Колонки автоматически НЕ
    создаются — метаданные лежат в values строк, колонки заводит агент
    через POST /columns. Дубликаты подсетей — skip."""
    if len(body.rows) > MAX_IMPORT_ROWS:
        raise HTTPException(400, f"За раз не больше {MAX_IMPORT_ROWS} строк "
                                 f"(получено {len(body.rows)})")
    try:
        res = store.import_flat_rows(provider_id, list_id, body.rows)
    except KeyError:
        raise HTTPException(404, "Список не найден")
    return {"ok": True, **res}


class OperatorBody(BaseModel):
    on: bool = True


@router.patch("/providers/{provider_id}/lists/{list_id}/rows/{row_id}/operator/{op_key}")
async def toggle_operator(provider_id: str, list_id: str, row_id: str, op_key: str, body: OperatorBody):
    try:
        store.toggle_operator(provider_id, list_id, row_id, op_key, body.on)
    except KeyError as e:
        raise _not_found(e)
    except ValueError as e:
        raise HTTPException(422, str(e))
    return {"ok": True}


class ColumnBody(BaseModel):
    title: str = ""
    # Кастомный ключ колонки (совпадает с ключом поля в row.values) —
    # например "operator"/"asn"/"country". Пусто → генерируется col_xxx.
    key: str = ""


@router.post("/providers/{provider_id}/lists/{list_id}/columns", status_code=201)
async def add_column(provider_id: str, list_id: str, body: ColumnBody):
    try:
        return store.add_column(provider_id, list_id, body.title, key=body.key)
    except KeyError as e:
        raise _not_found(e)


@router.patch("/providers/{provider_id}/lists/{list_id}/columns/{col_key}")
async def rename_column(provider_id: str, list_id: str, col_key: str, body: ColumnBody):
    try:
        store.rename_column(provider_id, list_id, col_key, body.title)
    except KeyError as e:
        raise _not_found(e)
    return {"ok": True}


@router.delete("/providers/{provider_id}/lists/{list_id}/columns/{col_key}", status_code=204)
async def delete_column(provider_id: str, list_id: str, col_key: str):
    try:
        store.delete_column(provider_id, list_id, col_key)
    except KeyError as e:
        raise _not_found(e)
    except ValueError as e:
        raise HTTPException(422, str(e))


class ColOrderBody(BaseModel):
    order: list[str] = []


@router.put("/providers/{provider_id}/lists/{list_id}/columns-order")
async def reorder_columns(provider_id: str, list_id: str, body: ColOrderBody):
    try:
        store.reorder_columns(provider_id, list_id, body.order)
    except KeyError as e:
        raise _not_found(e)
    return {"ok": True}


# ── обогащение ASN/провайдера (ip-api, как в «Анализе подписки») ─
IP_API_FIELDS = "status,as,asname,org,country"
# ip-api без ключа: 45 запросов/мин. Неразмеченные строки обогащаем пачками
# по 40 с паузой между пачками, чтобы не словить 429.
ENRICH_MISSING_LIMIT = 1000
ENRICH_BATCH_SIZE = 40
ENRICH_BATCH_SLEEP = 1.8


class EnrichBody(BaseModel):
    row_ids: list[str] = Field(..., min_length=1)


@router.post("/providers/{provider_id}/lists/{list_id}/enrich")
async def enrich_rows(provider_id: str, list_id: str, body: EnrichBody):
    """Обогащение строк по подсети (ip-api по сетевому адресу): заполняет
    asn/asnname/provider/country. Уже заполненные поля НЕ перезаписываются."""
    from app.services.subscription_analyze import _parse_as_field

    data = store.get_store()
    lst = next((l for p in data["providers"] if p["id"] == provider_id
                for l in p.get("lists", []) if l["id"] == list_id), None)
    if not lst:
        raise HTTPException(404, "Список не найден")
    rows = [r for r in lst.get("rows", []) if r.get("id") in set(body.row_ids)]
    if not rows:
        raise HTTPException(404, "Строки не найдены")

    updated = 0
    async with httpx.AsyncClient(timeout=8) as client:
        for row in rows:
            subnet = row.get("values", {}).get("subnet", "")
            ip = subnet.split("/")[0]
            try:
                r = await client.get(
                    f"http://ip-api.com/json/{ip}",
                    params={"fields": IP_API_FIELDS},
                )
                d = r.json()
            except Exception:
                continue
            if not isinstance(d, dict) or d.get("status") != "success":
                continue
            num, name = _parse_as_field(str(d.get("as") or ""))
            asnname = str(d.get("asname") or name or d.get("org") or "")
            org = str(d.get("org") or "")
            store.update_row_asn(provider_id, list_id, row["id"],
                                 f"AS{num}" if num else "", asnname,
                                 provider=org or asnname,
                                 country=str(d.get("country") or ""))
            updated += 1
    return {"updated": updated, "of": len(rows)}


class EnrichMissingBody(BaseModel):
    # Поля, по которым ищутся «неразмеченные» строки. Пусто → provider.
    fields: list[str] = Field(default_factory=list)


@router.post("/providers/{provider_id}/lists/{list_id}/enrich-missing")
async def enrich_missing(provider_id: str, list_id: str,
                         body: EnrichMissingBody | None = None):
    """Обогатить ВСЕ неразмеченные строки списка одним действием.

    Кандидаты — строки, у которых пусто хотя бы одно из body.fields
    (по умолчанию provider). Заполняются пустые asn/asnname/provider/country;
    уже заполненные поля не трогаются. Из-за лимита ip-api (45 req/min) —
    пачками по ENRICH_BATCH_SIZE с паузой между ними; за вызов не больше
    ENRICH_MISSING_LIMIT строк. Ответ: {updated, of, skipped}."""
    from app.services.subscription_analyze import _parse_as_field

    data = store.get_store()
    lst = next((l for p in data["providers"] if p["id"] == provider_id
                for l in p.get("lists", []) if l["id"] == list_id), None)
    if not lst:
        raise HTTPException(404, "Список не найден")
    wanted = (body.fields if body and body.fields else ["provider"])

    def _missing(row: dict) -> bool:
        values = row.get("values") or {}
        return any(not str(values.get(f) or "").strip() for f in wanted)

    targets = [r for r in lst.get("rows", []) if _missing(r)]
    of = len(targets)
    targets = targets[:ENRICH_MISSING_LIMIT]

    updated = 0
    async with httpx.AsyncClient(timeout=8) as client:
        for i in range(0, len(targets), ENRICH_BATCH_SIZE):
            batch = targets[i:i + ENRICH_BATCH_SIZE]
            for row in batch:
                subnet = row.get("values", {}).get("subnet", "")
                ip = subnet.split("/")[0]
                try:
                    r = await client.get(
                        f"http://ip-api.com/json/{ip}",
                        params={"fields": IP_API_FIELDS},
                    )
                    d = r.json()
                except Exception:
                    continue
                if not isinstance(d, dict) or d.get("status") != "success":
                    continue
                num, name = _parse_as_field(str(d.get("as") or ""))
                asnname = str(d.get("asname") or name or d.get("org") or "")
                org = str(d.get("org") or "")
                store.update_row_asn(provider_id, list_id, row["id"],
                                     f"AS{num}" if num else "", asnname,
                                     provider=org or asnname,
                                     country=str(d.get("country") or ""))
                updated += 1
            if i + ENRICH_BATCH_SIZE < len(targets):
                await asyncio.sleep(ENRICH_BATCH_SLEEP)
    return {"updated": updated, "of": of, "skipped": of - updated}


class EnrichTypesBody(BaseModel):
    # Пусто = все строки списка.
    row_ids: list[str] = Field(default_factory=list)


@router.post("/providers/{provider_id}/lists/{list_id}/enrich-types")
async def enrich_types(provider_id: str, list_id: str,
                       body: EnrichTypesBody | None = None):
    """Тип ASN (isp/hosting/business) по ТЕКУЩИМ данным строки — эвристика
    по org/asnname/netname/provider, ip-api НЕ вызывается, provider не
    меняется. Колонка asn_type создаётся, если её ещё нет, и заполняется
    у строк с непустыми данными. Ответ: {updated, of}."""
    data = store.get_store()
    lst = next((l for p in data["providers"] if p["id"] == provider_id
                for l in p.get("lists", []) if l["id"] == list_id), None)
    if not lst:
        raise HTTPException(404, "Список не найден")
    rows = lst.get("rows", [])
    if body and body.row_ids:
        wanted = set(body.row_ids)
        rows = [r for r in rows if r.get("id") in wanted]
    if not rows:
        raise HTTPException(404, "Строки не найдены")

    if "asn_type" not in {c.get("key") for c in lst.get("columns", [])}:
        store.add_column(provider_id, list_id, "Тип ASN", key="asn_type")

    updated = 0
    for row in rows:
        values = row.get("values") or {}
        t = store._asn_type(str(values.get("org") or ""),
                            str(values.get("asnname") or ""),
                            str(values.get("netname") or ""),
                            str(values.get("provider") or ""))
        if not t or str(values.get("asn_type") or "") == t:
            continue  # данных нет или тип уже проставлен — не трогаем
        store.set_cell(provider_id, list_id, row["id"], "asn_type", t)
        updated += 1
    return {"updated": updated, "of": len(rows)}


# ── иконки провайдеров/списков (файлы в DATA_DIR, раздача через API) ──
@router.post("/providers/{provider_id}/icon")
async def upload_provider_icon(provider_id: str, file: UploadFile = File(...)):
    """Загрузить иконку провайдера (png/svg/webp ≤ 256 КБ, multipart)."""
    try:
        store.save_provider_icon(provider_id, await file.read(),
                                 file.filename or "")
    except KeyError as e:
        raise _not_found(e)
    except ValueError as e:
        raise HTTPException(400, str(e))
    return {"ok": True}


@router.post("/providers/{provider_id}/lists/{list_id}/icon")
async def upload_list_icon(provider_id: str, list_id: str,
                           file: UploadFile = File(...)):
    """Загрузить иконку списка (png/svg/webp ≤ 256 КБ, multipart)."""
    try:
        store.save_list_icon(provider_id, list_id, await file.read(),
                             file.filename or "")
    except KeyError as e:
        raise _not_found(e)
    except ValueError as e:
        raise HTTPException(400, str(e))
    return {"ok": True}


def _icon_response(provider_id: str, list_id: str = "") -> FileResponse:
    data = store.get_store()
    if list_id:
        lst = next((l for p in data["providers"] if p["id"] == provider_id
                    for l in p.get("lists", []) if l["id"] == list_id), None)
        if not lst:
            raise _not_found(KeyError(list_id))
        name = lst.get("icon") or ""
    else:
        p = next((x for x in data["providers"] if x.get("id") == provider_id), None)
        if not p:
            raise _not_found(KeyError(provider_id))
        name = p.get("icon") or ""
    path = store.icon_file(name)
    if path is None:
        raise HTTPException(404, "Иконка не загружена")
    return FileResponse(path)


@router.get("/provider-icon/{provider_id}")
async def get_provider_icon(provider_id: str):
    return _icon_response(provider_id)


@router.get("/list-icon/{provider_id}/{list_id}")
async def get_list_icon(provider_id: str, list_id: str):
    return _icon_response(provider_id, list_id)


# ── Latency Lab: замер подсетей списка ────────────────────────
#
# Потолок на пачку. Мультискан — ОДИН запрос суточного лимита независимо от
# числа целей, а вот поштучный `subnet-scan` тратит по запросу на подсеть,
# поэтому случайный «выделить всё» на списке в 1000+ строк обнулил бы лимит
# аккаунта одним нажатием. Фронт режет большие выборки на порции по 750.
MAX_SCAN_SUBNETS = 750


class LatencyScanBody(BaseModel):
    provider_id: str
    list_id: str
    row_ids: list[str] = Field(default_factory=list)
    all: bool = False
    #: Пусто = мультискан по всем online-операторам (1 запрос лимита);
    #: конкретный оператор = поштучный subnet-scan.
    operator: str = ""
    async_: bool = Field(False, alias="async_")

    model_config = {"populate_by_name": True}


def _collect_subnets(provider_id: str, list_id: str, row_ids: list[str],
                     take_all: bool) -> list[str]:
    """Подсети выбранных строк: колонка `subnet`, иначе первая колонка списка."""
    data = store.get_store()
    lst = next((l for p in data["providers"] if p["id"] == provider_id
                for l in p.get("lists", []) if l["id"] == list_id), None)
    if not lst:
        raise HTTPException(404, "Список не найден")
    rows = lst.get("rows", [])
    if not take_all:
        wanted = set(row_ids)
        rows = [r for r in rows if r.get("id") in wanted]
    if not rows:
        raise HTTPException(404, "Строки не найдены")
    keys = [c["key"] for c in lst.get("columns", [])] or ["subnet"]
    key = "subnet" if "subnet" in keys else keys[0]
    out: list[str] = []
    for row in rows:
        value = (row.get("values", {}).get(key) or "").strip()
        if value and value not in out:
            out.append(value)
    if not out:
        raise HTTPException(400, "В выбранных строках нет подсетей")
    return out


def _split_host_subnets(subnets: list[str]) -> tuple[list[str], list[str]]:
    """(сети, host-адреса). Мультискан Latency Lab не распознаёт host-адреса
    /32 (IPv4) и /128 (IPv6) — «не удалось распознать сеть»; поштучный
    subnet-scan их принимает (/23…/32). Мусор не роняет: strip, без пустых."""
    networks: list[str] = []
    hosts: list[str] = []
    for s in subnets:
        s = s.strip()
        if not s:
            continue
        try:
            ip_str, prefix = s.split("/", 1)
            is_v6 = ":" in ip_str
            if (is_v6 and int(prefix) >= 128) or (not is_v6 and int(prefix) >= 32):
                hosts.append(s)
            else:
                networks.append(s)
        except ValueError:
            networks.append(s)  # мусор — пусть разбирается сам сервис
    return networks, hosts


def _latency_client():
    """Клиент Latency Lab или 400: ключ обязателен и интеграция включена."""
    cfg = latency_lab.config()
    if not cfg.enabled:
        raise HTTPException(400, "Latency Lab выключен в настройках")
    if not cfg.api_key_enc:
        raise HTTPException(400, "Не задан API-ключ Latency Lab")
    client = latency_lab.client()
    if client is None:
        raise HTTPException(400, "Не удалось расшифровать API-ключ Latency Lab")
    return client


@router.post("/latency-scan")
async def latency_scan(body: LatencyScanBody):
    """Замер выбранных подсетей через Latency Lab.

    Без оператора — мультискан (все online-операторы, один запрос лимита);
    с оператором — поштучный `subnet-scan` по каждой подсети.
    """
    client = _latency_client()
    cfg = latency_lab.config()
    used, limit = latency_lab.scan_quota(cfg)
    if limit and used >= limit:
        raise HTTPException(
            429, f"Превышен лимит сканов: {limit} за "
                 f"{cfg.scan_window_hours} ч. Подождите или измените лимит "
                 f"в настройках Latency Lab")
    subnets = _collect_subnets(body.provider_id, body.list_id,
                               body.row_ids, body.all)
    if len(subnets) > MAX_SCAN_SUBNETS:
        raise HTTPException(
            400, f"За раз не больше {MAX_SCAN_SUBNETS} подсетей "
                 f"(выбрано {len(subnets)})")

    operator = latency_lab.normalize_operator(body.operator or cfg.default_operator)
    if operator and operator not in latency_lab.OPERATORS:
        raise HTTPException(400, f"Неизвестный оператор: {operator}")

    jobs: list[dict] = []
    errors: list[str] = []

    if not operator:
        # Мультискан принимает цели одним текстом — это ровно один запрос
        # суточного лимита на всю пачку. Host-адреса /32 (IPv6 — /128)
        # мультискан НЕ распознаёт («не удалось распознать сеть»): их умеет
        # только поштучный subnet-scan, а он требует оператора. Без оператора
        # — понятная ошибка на каждый host-адрес, сети сканируются как раньше.
        networks, hosts = _split_host_subnets(subnets)
        if hosts and not networks:
            raise HTTPException(
                400, "host-адреса не поддерживаются мультисканом: "
                     f"{'; '.join(hosts[:5])}{'…' if len(hosts) > 5 else ''}"
                     " — укажите оператора для поштучного скана")
        for h in hosts:
            errors.append(f"{h}: host-адрес /32 не поддерживается мультисканом"
                          " — укажите оператора для поштучного скана")
        if networks:
            data, err = await client.multiscan("\n".join(networks),
                                               is_async=body.async_)
            if err:
                raise HTTPException(502, f"Latency Lab: {err}")
            data = data or {}
            jobs.append({"targets": networks, "req_id": data.get("req_id", ""),
                         "status": data.get("status", "done"),
                         "result": data.get("result")})
    else:
        for subnet in subnets:
            data, err = await client.subnet_scan(operator, subnet,
                                                 is_async=body.async_)
            if err:
                errors.append(f"{subnet}: {err}")
                continue
            data = data or {}
            jobs.append({"targets": [subnet], "req_id": data.get("req_id", ""),
                         "status": data.get("status", "done"),
                         "result": data.get("result")})
        if not jobs:
            raise HTTPException(502, "Latency Lab: " + "; ".join(errors[:3]))

    # Лимит «N сканов за M часов»: метка пишется ТОЛЬКО когда Latency Lab
    # принял скан (дошли сюда без исключения) — ошибка сервиса лимит не тратит.
    latency_lab.record_scan()

    return {"ok": True,
            "mode": "multiscan" if not operator else "subnet-scan",
            "operator": operator, "async": body.async_,
            "jobs": jobs, "errors": errors}


@router.get("/latency-scan/{req_id}")
async def latency_scan_status(req_id: str):
    client = _latency_client()
    data, err = await client.job_status(req_id)
    if err:
        return {"ok": False, "error": err, "req_id": req_id}
    data = data or {}
    return {"ok": True, "req_id": req_id,
            # Статусы-синонимы Latency Lab («success»/«failed») приводим к
            # каноническим done/error — иначе поллинг ждёт «done» вечно.
            "status": latency_lab.normalize_job_status(data.get("status", "")),
            "result": data.get("result")}


class LatencyCancelBody(BaseModel):
    req_id: str


@router.post("/latency-scan/cancel")
async def latency_scan_cancel(body: LatencyCancelBody):
    client = _latency_client()
    data, err = await client.cancel(body.req_id)
    if err:
        return {"ok": False, "error": err, "req_id": body.req_id}
    return {"ok": True, "req_id": body.req_id,
            "result": (data or {}).get("result")}


# ── импорт/экспорт (json/csv/txt) ─────────────────────────────
#
# Плоские форматы (csv/txt) описывают ровно один список, поэтому требуют
# list_id; json умеет и полный снимок дерева, и снимок одного списка.
MAX_IMPORT_ROWS = 5000
JSON_FORMAT = "na-subnets"
JSON_VERSION = 1
_DEFAULT_KEYS = {c["key"] for c in store.DEFAULT_COLUMNS}
_OP_KEYS = [o["key"] for o in store.OPERATORS]
_CSV_ALIASES = {"version": "ipver", "версия ip": "ipver", "подсеть": "subnet",
                "дата": "date", "название asn": "asnname"}
_MEDIA = {"json": "application/json; charset=utf-8",
          "csv": "text/csv; charset=utf-8",
          "txt": "text/plain; charset=utf-8",
          "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}


def _pick_list(provider_id: str, list_id: str) -> tuple[dict, dict]:
    data = store.get_store()
    for p in data["providers"]:
        for l in p.get("lists", []):
            if l.get("id") == list_id and (not provider_id or p.get("id") == provider_id):
                return p, l
    raise HTTPException(404, "Список не найден")


def _flat_header(lst: dict) -> list[tuple[str, str]]:
    """Колонки плоского CSV: (заголовок, ключ). operators разворачивается."""
    out: list[tuple[str, str]] = []
    for col in lst.get("columns", []):
        key = col.get("key") or ""
        if key == "operators":
            out += [(k, f"op:{k}") for k in _OP_KEYS]
        elif key in _DEFAULT_KEYS:
            out.append((key, key))
        else:
            out.append((col.get("title") or key, key))
    return out


def _list_to_csv(lst: dict) -> str:
    header = _flat_header(lst)
    buf = io.StringIO()
    w = csv.writer(buf, lineterminator="\n")
    w.writerow([h for h, _ in header])
    for row in lst.get("rows", []):
        values, ops = row.get("values") or {}, row.get("operators") or {}
        line = []
        for _, key in header:
            if key.startswith("op:"):
                line.append("1" if ops.get(key[3:], False) else "0")
            else:
                line.append(str(values.get(key) or ""))
        w.writerow(line)
    return buf.getvalue()


def _list_to_txt(lst: dict) -> str:
    out = [f"# {lst.get('name') or 'Список'}"]
    for row in lst.get("rows", []):
        subnet = (row.get("values") or {}).get("subnet") or ""
        if subnet:
            out.append(subnet)
    return "\n".join(out) + "\n"


def _prov_key(row: dict) -> tuple:
    """Ключ сортировки/группировки по провайдеру: без провайдера — в конец."""
    prov = ((row.get("values") or {}).get("provider") or "").strip()
    return (1 if not prov else 0, prov)


def _list_to_xlsx(lst: dict) -> bytes:
    """Список → .xlsx (bytes). Колонки как в CSV (operators → 5 колонок,
    значения 1/0); строки отсортированы по провайдеру (пустые — в конец,
    группа «—») и сгруппированы: ws.row_dimensions.group(start, end,
    outline_level=1) — сворачивание «+/-» в Excel, hidden=False.
    Заголовок жирный, первая строка закреплена (freeze_panes="A2"),
    автоширина колонок (cap 40)."""
    header = _flat_header(lst)
    op_label = {o["key"]: o["label"] for o in store.OPERATORS}
    titles = [op_label.get(k[3:], h) if k.startswith("op:") else h
              for h, k in header]

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    name = "".join(c if c not in "[]:*?/\\" else " " for c in (lst.get("name") or ""))
    ws.title = (name.strip() or "Таблица")[:31]
    ws.append(titles)

    data = sorted(lst.get("rows") or [], key=_prov_key)
    # Группировка по провайдеру: заголовок группы (уровень 0) + строки
    # данных (уровень 1). Без разрыва уровня Excel сливает все строки в
    # одну гигантскую группу; заголовок уровня 0 между группами даёт
    # отдельные «+/-» по каждому провайдеру.
    i = 0
    while i < len(data):
        key = _prov_key(data[i])
        prov = key[1] or "—"
        # строка-заголовок группы: имя провайдера в первой колонке
        grp = i
        while grp < len(data) and _prov_key(data[grp]) == key:
            grp += 1
        ws.append([f"{prov} ({grp - i})"] + [""] * (len(header) - 1))
        hdr_row = ws.max_row
        for c in ws[hdr_row]:
            c.font = Font(bold=True)
        start = ws.max_row + 1
        for row in data[i:grp]:
            values, ops = row.get("values") or {}, row.get("operators") or {}
            line = []
            for _, key2 in header:
                if key2.startswith("op:"):
                    line.append(1 if ops.get(key2[3:], False) else 0)
                else:
                    line.append(values.get(key2) or "")
            ws.append(line)
        end = ws.max_row
        if start <= end:
            ws.row_dimensions.group(start, end, outline_level=1)
        i = grp

    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for col in ws.columns:
        letter = col[0].column_letter
        width = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[letter].width = min(max(width + 2, 10), 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/export")
async def export_subnets(provider_id: str = "", list_id: str = "", format: str = "json"):
    fmt = (format or "json").strip().lower()
    if fmt not in _MEDIA:
        raise HTTPException(400, "Формат должен быть json, csv, txt или xlsx")
    if fmt in ("csv", "txt", "xlsx") and not list_id:
        raise HTTPException(400, "Для формата csv/txt/xlsx укажите список (list_id)")

    if list_id:
        provider, lst = _pick_list(provider_id, list_id)
    else:
        provider, lst = None, None

    if fmt == "json":
        if lst is not None and provider is not None:
            payload = {"_format": JSON_FORMAT, "_version": JSON_VERSION,
                       "providers": [{"id": provider["id"], "name": provider["name"],
                                      "lists": [lst]}]}
        else:
            data = store.get_store()
            providers = data["providers"]
            if provider_id:
                providers = [p for p in providers if p.get("id") == provider_id]
                if not providers:
                    raise HTTPException(404, "Провайдер не найден")
            payload = {"_format": JSON_FORMAT, "_version": JSON_VERSION,
                       "providers": providers}
        body = json.dumps(payload, ensure_ascii=False, indent=1)
    elif fmt == "csv":
        body = _list_to_csv(lst or {})
    elif fmt == "txt":
        body = _list_to_txt(lst or {})
    else:
        body = _list_to_xlsx(lst or {})

    name = f"subnets_{time.strftime('%Y%m%d-%H%M%S')}.{fmt}"
    payload = body if isinstance(body, bytes) else body.encode("utf-8")
    return Response(payload, media_type=_MEDIA[fmt],
                    headers={"Content-Disposition": f'attachment; filename="{name}"'})


def _detect_format(filename: str, text: str) -> str:
    ext = (filename or "").rsplit(".", 1)[-1].lower() if "." in (filename or "") else ""
    if ext in ("json", "csv", "txt"):
        return ext
    head = text.lstrip()[:1]
    if head in ("{", "["):
        return "json"
    first = next((l for l in text.splitlines() if l.strip()), "")
    return "csv" if ("," in first or ";" in first) else "txt"


def _parse_txt(text: str) -> list[dict]:
    items = []
    for line in text.splitlines():
        s = line.strip()
        if not s or s.startswith("#") or s.startswith("//"):
            continue
        for sep in ("#", "—", " - ", "\t", ";", ","):
            if sep in s:
                s = s.split(sep, 1)[0].strip()
        if s:
            items.append({"subnet": s})
    return items


_TRUE = {"1", "true", "yes", "y", "да", "+", "on", "истина"}


def _parse_csv(text: str) -> list[dict]:
    sample = next((l for l in text.splitlines() if l.strip()), "")
    delim = ";" if sample.count(";") > sample.count(",") else ","
    rows = list(csv.reader(io.StringIO(text), delimiter=delim))
    rows = [r for r in rows if any((c or "").strip() for c in r)]
    if not rows:
        return []
    head = [(c or "").strip() for c in rows[0]]
    lowered = [h.lower() for h in head]
    has_header = any(_CSV_ALIASES.get(h, h) == "subnet" for h in lowered)
    if not has_header:
        return [{"subnet": (r[0] or "").strip()} for r in rows if (r[0] or "").strip()]

    items = []
    for r in rows[1:]:
        cells = [(c or "").strip() for c in r]
        item: dict = {"subnet": "", "values": {}, "extra": {}, "operators": {}}
        for i, title in enumerate(head):
            val = cells[i] if i < len(cells) else ""
            key = _CSV_ALIASES.get(lowered[i], lowered[i])
            if key == "subnet":
                item["subnet"] = val
            elif key in _OP_KEYS:
                item["operators"][key] = val.strip().lower() in _TRUE
            elif key in _DEFAULT_KEYS:
                item["values"][key] = val
            elif title:
                item["extra"][title] = val
        if item["subnet"]:
            items.append(item)
    return items


def _snapshot_providers(payload) -> list:
    if isinstance(payload, list):
        return payload
    if not isinstance(payload, dict):
        raise HTTPException(400, "Некорректный JSON: ожидался объект снимка")
    providers = payload.get("providers")
    if isinstance(providers, list):
        return providers
    if isinstance(payload.get("lists"), list):  # снимок одного провайдера
        return [payload]
    if isinstance(payload.get("rows"), list):  # снимок одного списка
        return [{"name": "Импортированные", "lists": [payload]}]
    raise HTTPException(400, "Некорректный JSON: нет providers/lists/rows")


def _snapshot_items(providers: list) -> list[dict]:
    """Все строки снимка плоским списком (импорт в конкретный список)."""
    items: list[dict] = []
    for p in providers:
        if not isinstance(p, dict):
            continue
        for l in p.get("lists") or []:
            if isinstance(l, dict):
                items += store.rows_to_items(l.get("rows"))
    return items


@router.post("/import")
async def import_subnets(file: UploadFile = File(...), provider_id: str = Form(""),
                         list_id: str = Form(""), mode: str = Form("merge"),
                         format: str = Form("")):
    mode = (mode or "merge").strip().lower()
    if mode not in ("merge", "replace"):
        raise HTTPException(400, "Режим должен быть merge или replace")
    blob = await file.read()
    if not blob:
        raise HTTPException(400, "Файл пуст")
    try:
        text = blob.decode("utf-8-sig")
    except UnicodeDecodeError:
        raise HTTPException(400, "Файл должен быть в кодировке UTF-8")

    fmt = (format or "").strip().lower() or _detect_format(file.filename or "", text)
    if fmt not in ("json", "csv", "txt"):
        raise HTTPException(400, "Поддерживаются только json, csv и txt")

    if fmt == "json":
        try:
            payload = json.loads(text)
        except ValueError:
            raise HTTPException(400, "Некорректный JSON-файл")
        providers = _snapshot_providers(payload)
        rows_total = sum(len(l.get("rows") or [])
                         for p in providers if isinstance(p, dict)
                         for l in (p.get("lists") or []) if isinstance(l, dict))
        if rows_total > MAX_IMPORT_ROWS:
            raise HTTPException(400, f"За раз не больше {MAX_IMPORT_ROWS} строк "
                                     f"(в файле {rows_total})")
        if not list_id:
            res = store.import_tree(providers, replace=(mode == "replace"))
            return {"ok": True, **res}
        items = _snapshot_items(providers)
    else:
        items = _parse_csv(text) if fmt == "csv" else _parse_txt(text)
        if len(items) > MAX_IMPORT_ROWS:
            raise HTTPException(400, f"За раз не больше {MAX_IMPORT_ROWS} строк "
                                     f"(в файле {len(items)})")

    if not items:
        raise HTTPException(400, "В файле не найдено подсетей")

    if list_id:
        provider, lst = _pick_list(provider_id, list_id)
        pid, lid = provider["id"], lst["id"]
    else:
        provider = (store.ensure_provider("Импортированные") if not provider_id
                    else next((p for p in store.get_store()["providers"]
                               if p.get("id") == provider_id), None))
        if not provider:
            raise HTTPException(404, "Провайдер не найден")
        lst = store.ensure_list(provider["id"],
                                f"Импорт {time.strftime('%Y-%m-%d %H:%M:%S')}")
        pid, lid = provider["id"], lst["id"]

    try:
        res = store.import_rows(pid, lid, items, replace=(mode == "replace"))
    except KeyError:
        raise HTTPException(404, "Список не найден")
    return {"ok": True, **res}
