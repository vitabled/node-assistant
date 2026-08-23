"""Wave-5 PR-5 — «Подсети»: провайдеры/списки/строки/столбцы, обогащение,
импорт/экспорт (json/csv/txt/xlsx)."""
import io
import json as _json
import uuid as _uuid

import openpyxl
from fastapi.testclient import TestClient

from app.main import app
from app.services import subnets_store as store

client = TestClient(app)


def _auth():
    r = client.post("/api/auth/register",
                    json={"login": f"sn-{_uuid.uuid4().hex[:8]}", "password": "pw"})
    return {"Authorization": f"Bearer {r.json()['token']}"}


def _mk_list(a):
    p = client.post("/api/subnets/providers", headers=a, json={"name": "МТС"}).json()
    l = client.post(f"/api/subnets/providers/{p['id']}/lists", headers=a,
                    json={"name": "Основной"}).json()
    return p["id"], l["id"]


def test_provider_list_row_flow():
    a = _auth()
    pid, lid = _mk_list(a)
    r = client.post(f"/api/subnets/providers/{pid}/lists/{lid}/rows", headers=a,
                    json={"subnets": ["203.0.113.9/24", "2001:db8::/32", "мусор"]})
    assert r.status_code == 201
    assert r.json()["added"] == ["203.0.113.0/24", "2001:db8::/32"]
    assert len(r.json()["errors"]) == 1
    data = client.get("/api/subnets", headers=a).json()
    lst = data["providers"][0]["lists"][0]
    assert [c["key"] for c in lst["columns"]] == ["subnet", "ipver", "asn", "asnname", "date", "operators"]
    row = lst["rows"][0]
    assert row["values"]["subnet"] == "203.0.113.0/24"
    assert row["values"]["ipver"] == "IPv4"
    assert row["operators"]["mts"] is True
    assert data["operators"][1]["key"] == "beeline"


def test_parse_subnet():
    assert store.parse_subnet("1.2.3.4") == ("1.2.3.4/32", "IPv4")
    assert store.parse_subnet("1.2.3.9/24") == ("1.2.3.0/24", "IPv4")
    assert store.parse_subnet("2001:db8::1") == ("2001:db8::1/128", "IPv6")
    try:
        store.parse_subnet("not-an-ip")
        raise AssertionError("должно упасть")
    except ValueError:
        pass


def test_column_ops_and_operator_toggle():
    a = _auth()
    pid, lid = _mk_list(a)
    client.post(f"/api/subnets/providers/{pid}/lists/{lid}/rows", headers=a,
                json={"subnets": ["1.2.3.0/24"]})
    # новый столбец
    col = client.post(f"/api/subnets/providers/{pid}/lists/{lid}/columns", headers=a,
                      json={"title": "Комментарий"}).json()
    # порядок: комментарий первым
    r = client.put(f"/api/subnets/providers/{pid}/lists/{lid}/columns-order", headers=a,
                   json={"order": [col["key"], "subnet", "operators"]})
    assert r.status_code == 200
    data = client.get("/api/subnets", headers=a).json()
    cols = [c["key"] for c in data["providers"][0]["lists"][0]["columns"]]
    assert cols[0] == col["key"] and cols[1] == "subnet"
    # удаление «Подсети» запрещено
    r = client.delete(f"/api/subnets/providers/{pid}/lists/{lid}/columns/subnet", headers=a)
    assert r.status_code == 422
    # тоггл оператора
    rid = data["providers"][0]["lists"][0]["rows"][0]["id"]
    r = client.patch(f"/api/subnets/providers/{pid}/lists/{lid}/rows/{rid}/operator/beeline",
                     headers=a, json={"on": False})
    assert r.status_code == 200
    data = client.get("/api/subnets", headers=a).json()
    row = data["providers"][0]["lists"][0]["rows"][0]
    assert row["operators"]["beeline"] is False
    assert row["operators"]["mts"] is True


def test_enrich_marks_asn(monkeypatch):
    import app.api.subnets as api
    a = _auth()
    pid, lid = _mk_list(a)
    client.post(f"/api/subnets/providers/{pid}/lists/{lid}/rows", headers=a,
                json={"subnets": ["8.8.8.0/24"]})
    data = client.get("/api/subnets", headers=a).json()
    rid = data["providers"][0]["lists"][0]["rows"][0]["id"]

    class R:
        def json(self):
            return {"status": "success", "as": "AS15169 Google LLC", "asname": "Google", "org": "Google"}

    class FakeClient:
        async def __aenter__(self):
            return self

        async def __aexit__(self, *a):
            pass

        async def get(self, url, params=None):
            return R()

    monkeypatch.setattr(api.httpx, "AsyncClient", lambda *a, **k: FakeClient())
    r = client.post(f"/api/subnets/providers/{pid}/lists/{lid}/enrich", headers=a,
                    json={"row_ids": [rid]})
    assert r.status_code == 200 and r.json()["updated"] == 1
    data = client.get("/api/subnets", headers=a).json()
    row = data["providers"][0]["lists"][0]["rows"][0]
    assert row["values"]["asn"] == "AS15169"
    assert "Google" in row["values"]["asnname"]


class _EnrichResp:
    """Ответ ip-api: полный набор полей (org/country — для провайдера и страны)."""

    def json(self):
        return {"status": "success", "as": "AS15169 Google LLC", "asname": "Google",
                "org": "Google LLC", "country": "US"}


class _EnrichClient:
    """Фейк ip-api: журнал запросов (url, params) — для проверки пачек и fields."""

    calls: list = []

    def __init__(self, *a, **k):
        pass

    async def __aenter__(self):
        return self

    async def __aexit__(self, *a):
        pass

    async def get(self, url, params=None):
        _EnrichClient.calls.append((url, dict(params or {})))
        return _EnrichResp()


def _patch_ip_api(monkeypatch):
    import app.api.subnets as api
    _EnrichClient.calls = []
    monkeypatch.setattr(api.httpx, "AsyncClient", lambda *a, **k: _EnrichClient())
    return api


def test_enrich_rows_fills_provider_country_keeps_existing(monkeypatch):
    """Расширенный enrich_rows: ip-api просит country, пустые строки получают
    provider (из org) и country, уже заполненные поля НЕ перезаписываются."""
    a = _auth()
    pid, lid = _mk_list(a)
    client.post(f"/api/subnets/providers/{pid}/lists/{lid}/rows", headers=a,
                json={"rows": [
                    {"subnet": "8.8.8.0/24"},
                    {"subnet": "9.9.9.0/24", "provider": "RUVDS", "country": "DE",
                     "asn": "AS64500", "asnname": "Старое"},
                ]})
    by_subnet = {r["values"]["subnet"]: r for r in _rows(a)}
    api = _patch_ip_api(monkeypatch)
    r = client.post(f"/api/subnets/providers/{pid}/lists/{lid}/enrich", headers=a,
                    json={"row_ids": [by_subnet["8.8.8.0/24"]["id"],
                                      by_subnet["9.9.9.0/24"]["id"]]})
    assert r.status_code == 200 and r.json()["updated"] == 2
    # запрос ip-api просит и страну
    assert "country" in _EnrichClient.calls[0][1]["fields"].split(",")
    fresh = {x["values"]["subnet"]: x["values"] for x in _rows(a)}
    # пустая строка — заполнилась полностью
    assert fresh["8.8.8.0/24"]["provider"] == "Google LLC"
    assert fresh["8.8.8.0/24"]["country"] == "US"
    assert fresh["8.8.8.0/24"]["asn"] == "AS15169"
    # уже заполненная — НЕ перезаписана
    assert fresh["9.9.9.0/24"]["provider"] == "RUVDS"
    assert fresh["9.9.9.0/24"]["country"] == "DE"
    assert fresh["9.9.9.0/24"]["asn"] == "AS64500"
    assert fresh["9.9.9.0/24"]["asnname"] == "Старое"


def test_enrich_missing_fills_unmarked_and_keeps_marked(monkeypatch):
    """enrich-missing: обогащает ВСЕ строки без провайдера ({updated, of,
    skipped}), размеченная строка не тронута, запросы уходят только по
    неразмеченным."""
    a = _auth()
    pid, lid = _mk_list(a)
    client.post(f"/api/subnets/providers/{pid}/lists/{lid}/rows", headers=a,
                json={"rows": [
                    {"subnet": "8.8.8.0/24"},
                    {"subnet": "9.9.9.0/24"},
                    {"subnet": "10.0.0.0/24", "provider": "RUVDS"},
                ]})
    _patch_ip_api(monkeypatch)
    r = client.post(f"/api/subnets/providers/{pid}/lists/{lid}/enrich-missing", headers=a)
    assert r.status_code == 200
    assert r.json() == {"updated": 2, "of": 2, "skipped": 0}
    fresh = {x["values"]["subnet"]: x["values"] for x in _rows(a)}
    assert fresh["8.8.8.0/24"]["provider"] == "Google LLC"
    assert fresh["9.9.9.0/24"]["country"] == "US"
    assert fresh["10.0.0.0/24"]["provider"] == "RUVDS"  # размеченная не тронута
    assert len(_EnrichClient.calls) == 2  # ровно по неразмеченным


def test_enrich_missing_batches_with_sleep(monkeypatch):
    """45 строк без провайдера → пачки 40 + 5: 45 запросов, одна пауза
    ENRICH_BATCH_SLEEP между пачками (не словить 429 ip-api)."""
    a = _auth()
    pid, lid = _mk_list(a)
    subnets = [f"10.{i // 256}.{i % 256}.0/24" for i in range(45)]
    client.post(f"/api/subnets/providers/{pid}/lists/{lid}/rows", headers=a,
                json={"subnets": subnets})
    api = _patch_ip_api(monkeypatch)
    sleeps: list[float] = []

    async def _sleep(secs):
        sleeps.append(secs)

    monkeypatch.setattr(api.asyncio, "sleep", _sleep)
    r = client.post(f"/api/subnets/providers/{pid}/lists/{lid}/enrich-missing", headers=a)
    assert r.status_code == 200
    assert r.json()["updated"] == 45
    assert len(_EnrichClient.calls) == 45
    assert sleeps == [api.ENRICH_BATCH_SLEEP]


def test_enrich_missing_limit_1000(monkeypatch):
    """Лимит 1000 строк за вызов: 1001 без провайдера → обработаны 1000,
    ответ {updated: 1000, of: 1001, skipped: 1}, за лимит не ходили."""
    a = _auth()
    pid, lid = _mk_list(a)
    subnets = [f"10.{i // 256}.{i % 256}.0/24" for i in range(1001)]
    client.post(f"/api/subnets/providers/{pid}/lists/{lid}/rows", headers=a,
                json={"subnets": subnets})
    api = _patch_ip_api(monkeypatch)
    sleeps: list[float] = []

    async def _sleep(secs):
        sleeps.append(secs)

    monkeypatch.setattr(api.asyncio, "sleep", _sleep)
    r = client.post(f"/api/subnets/providers/{pid}/lists/{lid}/enrich-missing", headers=a)
    assert r.status_code == 200
    body = r.json()
    assert body["updated"] == 1000 and body["of"] == 1001 and body["skipped"] == 1
    assert len(_EnrichClient.calls) == 1000
    assert len(sleeps) == 24  # 1000/40 = 25 пачек → 24 паузы


# ── импорт/экспорт (json/csv/txt) ─────────────────────────────
def _rows(a, pid=0, lid=0):
    data = client.get("/api/subnets", headers=a).json()
    return data["providers"][pid]["lists"][lid]["rows"]


def test_export_json_tree():
    a = _auth()
    pid, lid = _mk_list(a)
    client.post(f"/api/subnets/providers/{pid}/lists/{lid}/rows", headers=a,
                json={"subnets": ["203.0.113.0/24", "198.51.100.0/24"]})
    r = client.get("/api/subnets/export", headers=a)
    assert r.status_code == 200
    assert "attachment" in r.headers["content-disposition"]
    assert ".json" in r.headers["content-disposition"]
    body = r.json()
    assert body["_format"] == "na-subnets" and body["_version"] == 1
    assert len(body["providers"][0]["lists"][0]["rows"]) == 2


def test_export_csv_list():
    a = _auth()
    pid, lid = _mk_list(a)
    client.post(f"/api/subnets/providers/{pid}/lists/{lid}/rows", headers=a,
                json={"subnets": ["203.0.113.0/24"]})
    r = client.get("/api/subnets/export", headers=a,
                   params={"provider_id": pid, "list_id": lid, "format": "csv"})
    assert r.status_code == 200
    assert r.headers["content-type"].startswith("text/csv")
    lines = r.text.strip().splitlines()
    assert lines[0].split(",")[0] == "subnet"
    assert "mts" in lines[0]
    assert lines[1].startswith("203.0.113.0/24,IPv4")
    # csv без списка — 400
    assert client.get("/api/subnets/export", headers=a,
                      params={"format": "csv"}).status_code == 400
    # неизвестный формат — 400
    assert client.get("/api/subnets/export", headers=a,
                      params={"format": "xml", "list_id": lid}).status_code == 400


def test_export_txt_list():
    a = _auth()
    pid, lid = _mk_list(a)
    client.post(f"/api/subnets/providers/{pid}/lists/{lid}/rows", headers=a,
                json={"subnets": ["203.0.113.0/24", "2001:db8::/32"]})
    r = client.get("/api/subnets/export", headers=a,
                   params={"list_id": lid, "format": "txt"})
    assert r.status_code == 200
    lines = [l for l in r.text.splitlines() if l and not l.startswith("#")]
    assert lines == ["203.0.113.0/24", "2001:db8::/32"]


def test_import_json_merge_skips_duplicates():
    a = _auth()
    pid, lid = _mk_list(a)
    client.post(f"/api/subnets/providers/{pid}/lists/{lid}/rows", headers=a,
                json={"subnets": ["203.0.113.0/24"]})
    snapshot = client.get("/api/subnets/export", headers=a).json()
    rows = snapshot["providers"][0]["lists"][0]["rows"]
    rows.append({"id": "x", "values": {"subnet": "198.51.100.0/24"}, "operators": {}})
    blob = _json.dumps(snapshot).encode()
    r = client.post("/api/subnets/import", headers=a,
                    files={"file": ("snap.json", blob, "application/json")},
                    data={"mode": "merge"})
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["ok"] is True and body["imported"] == 1 and body["skipped"] == 1
    subnets = {r_["values"]["subnet"] for r_ in _rows(a)}
    assert subnets == {"203.0.113.0/24", "198.51.100.0/24"}


def test_import_json_replace_into_list():
    a = _auth()
    pid, lid = _mk_list(a)
    client.post(f"/api/subnets/providers/{pid}/lists/{lid}/rows", headers=a,
                json={"subnets": ["203.0.113.0/24"]})
    snap = {"_format": "na-subnets", "_version": 1, "providers": [
        {"name": "X", "lists": [{"name": "Y", "rows": [
            {"values": {"subnet": "10.0.0.0/8"}, "operators": {"mts": False}}]}]}]}
    r = client.post("/api/subnets/import", headers=a,
                    files={"file": ("s.json", _json.dumps(snap).encode(), "application/json")},
                    data={"provider_id": pid, "list_id": lid, "mode": "replace"})
    assert r.status_code == 200 and r.json()["imported"] == 1
    rows = _rows(a)
    assert len(rows) == 1 and rows[0]["values"]["subnet"] == "10.0.0.0/8"
    assert rows[0]["operators"]["mts"] is False


def test_import_csv_with_header_sets_operators():
    a = _auth()
    pid, lid = _mk_list(a)
    csv_text = ("subnet,version,asn,asnname,date,mts,beeline,Комментарий\n"
                "203.0.113.5/24,IPv4,AS64500,Example,2026-01-01,1,0,тест\n"
                "мусор,IPv4,,,,,,\n")
    r = client.post("/api/subnets/import", headers=a,
                    files={"file": ("rows.csv", csv_text.encode(), "text/csv")},
                    data={"provider_id": pid, "list_id": lid})
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["imported"] == 1 and body["skipped"] == 1 and body["errors"]
    row = _rows(a)[0]
    assert row["values"]["subnet"] == "203.0.113.0/24"
    assert row["values"]["asn"] == "AS64500"
    assert row["operators"]["mts"] is True and row["operators"]["beeline"] is False
    lst = client.get("/api/subnets", headers=a).json()["providers"][0]["lists"][0]
    ckey = next(c["key"] for c in lst["columns"] if c["title"] == "Комментарий")
    assert row["values"][ckey] == "тест"


def test_import_txt_ignores_comments():
    a = _auth()
    pid, lid = _mk_list(a)
    txt = ("# заголовок\n\n203.0.113.0/24\n"
           "198.51.100.0/24 — офис\n2001:db8::/32 # ipv6\n")
    r = client.post("/api/subnets/import", headers=a,
                    files={"file": ("list.txt", txt.encode(), "text/plain")},
                    data={"provider_id": pid, "list_id": lid})
    assert r.status_code == 200, r.text
    assert r.json()["imported"] == 3
    assert {x["values"]["subnet"] for x in _rows(a)} == {
        "203.0.113.0/24", "198.51.100.0/24", "2001:db8::/32"}


def test_import_creates_new_list():
    a = _auth()
    r = client.post("/api/subnets/import", headers=a,
                    files={"file": ("l.txt", b"203.0.113.0/24\n", "text/plain")})
    assert r.status_code == 200 and r.json()["imported"] == 1
    data = client.get("/api/subnets", headers=a).json()
    assert data["providers"][0]["name"] == "Импортированные"
    lst = data["providers"][0]["lists"][0]
    assert lst["name"].startswith("Импорт ") and len(lst["rows"]) == 1


def test_import_invalid_file_400():
    a = _auth()
    pid, lid = _mk_list(a)
    # битый json
    r = client.post("/api/subnets/import", headers=a,
                    files={"file": ("b.json", b"{not json", "application/json")},
                    data={"provider_id": pid, "list_id": lid})
    assert r.status_code == 400 and "JSON" in r.json()["detail"]
    # пустой файл
    r = client.post("/api/subnets/import", headers=a,
                    files={"file": ("e.txt", b"", "text/plain")})
    assert r.status_code == 400
    # без подсетей
    r = client.post("/api/subnets/import", headers=a,
                    files={"file": ("c.txt", "# только комментарий\n".encode(),
                                    "text/plain")})
    assert r.status_code == 400
    # неизвестный режим
    r = client.post("/api/subnets/import", headers=a,
                    files={"file": ("l.txt", b"1.2.3.0/24\n", "text/plain")},
                    data={"mode": "bogus"})
    assert r.status_code == 400
    # несуществующий список
    r = client.post("/api/subnets/import", headers=a,
                    files={"file": ("l.txt", b"1.2.3.0/24\n", "text/plain")},
                    data={"list_id": "nope"})
    assert r.status_code == 404


def test_import_limit_5000_rows_400():
    a = _auth()
    pid, lid = _mk_list(a)
    import app.api.subnets as api
    big = "\n".join(f"10.{i // 256}.{i % 256}.0/24" for i in range(api.MAX_IMPORT_ROWS + 1))
    r = client.post("/api/subnets/import", headers=a,
                    files={"file": ("big.txt", big.encode(), "text/plain")},
                    data={"provider_id": pid, "list_id": lid})
    assert r.status_code == 400 and "5000" in r.json()["detail"]


# ── rows с метаданными / batch-ячейки / import-json ────────────
def test_rows_with_metadata():
    a = _auth()
    pid, lid = _mk_list(a)
    r = client.post(f"/api/subnets/providers/{pid}/lists/{lid}/rows", headers=a,
                    json={"rows": [
                        {"subnet": "203.0.113.9/24", "operator": "mts",
                         "country": "RU", "asn": "AS64500"},
                        {"subnet": "198.51.100.0/24", "operators": {"beeline": False},
                         "comment": "офис"},
                    ]})
    assert r.status_code == 201, r.text
    body = r.json()
    assert body["added"] == ["203.0.113.0/24", "198.51.100.0/24"]
    assert body["errors"] == []
    rows = _rows(a)
    row = next(x for x in rows if x["values"]["subnet"] == "203.0.113.0/24")
    assert row["values"]["country"] == "RU"
    assert row["values"]["asn"] == "AS64500"
    assert row["operators"]["mts"] is True
    row2 = next(x for x in rows if x["values"]["subnet"] == "198.51.100.0/24")
    assert row2["operators"]["beeline"] is False
    assert row2["values"]["comment"] == "офис"
    # колонки автоматически не создаются
    lst = client.get("/api/subnets", headers=a).json()["providers"][0]["lists"][0]
    assert [c["key"] for c in lst["columns"]] == ["subnet", "ipver", "asn", "asnname", "date", "operators"]


def test_rows_old_format_still_works():
    a = _auth()
    pid, lid = _mk_list(a)
    r = client.post(f"/api/subnets/providers/{pid}/lists/{lid}/rows", headers=a,
                    json={"subnets": ["203.0.113.9/24", "2001:db8::/32", "мусор"]})
    assert r.status_code == 201
    body = r.json()
    assert body["added"] == ["203.0.113.0/24", "2001:db8::/32"]
    assert len(body["errors"]) == 1
    row = _rows(a)[0]
    assert row["values"]["subnet"] == "203.0.113.0/24"
    assert row["values"]["ipver"] == "IPv4"


def test_rows_body_requires_subnets_or_rows():
    a = _auth()
    pid, lid = _mk_list(a)
    r = client.post(f"/api/subnets/providers/{pid}/lists/{lid}/rows", headers=a, json={})
    assert r.status_code == 422
    r = client.post(f"/api/subnets/providers/{pid}/lists/{lid}/rows", headers=a,
                    json={"subnets": [], "rows": []})
    assert r.status_code == 422


def test_rows_unknown_operator_kept_as_is():
    a = _auth()
    pid, lid = _mk_list(a)
    r = client.post(f"/api/subnets/providers/{pid}/lists/{lid}/rows", headers=a,
                    json={"rows": [{"subnet": "203.0.113.0/24", "operator": "yota"}]})
    assert r.status_code == 201
    row = _rows(a)[0]
    assert row["values"]["operator"] == "yota"  # как есть, без падения
    assert row["operators"]["mts"] is False  # неизвестный оператор → все флаги False
    # битая строка в пачке не убивает остальные
    r = client.post(f"/api/subnets/providers/{pid}/lists/{lid}/rows", headers=a,
                    json={"rows": [{"subnet": "мусор", "country": "XX"},
                                   {"subnet": "198.51.100.0/24"}]})
    assert r.status_code == 201
    body = r.json()
    assert body["added"] == ["198.51.100.0/24"] and len(body["errors"]) == 1


def test_batch_update_cells():
    a = _auth()
    pid, lid = _mk_list(a)
    client.post(f"/api/subnets/providers/{pid}/lists/{lid}/rows", headers=a,
                json={"subnets": ["203.0.113.0/24", "198.51.100.0/24", "10.0.0.0/8"]})
    rows = _rows(a)
    r = client.patch(f"/api/subnets/providers/{pid}/lists/{lid}/rows/batch", headers=a,
                     json={"updates": [
                         {"row_id": rows[0]["id"], "col": "asn", "value": "AS64500"},
                         {"row_id": rows[1]["id"], "col": "asnname", "value": "Example"},
                         {"row_id": rows[2]["id"], "col": "date", "value": "2026-08-23"},
                     ]})
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["ok"] is True and body["updated"] == 3 and body["skipped"] == 0
    assert body["errors"] == []
    fresh = _rows(a)
    by_id = {x["id"]: x for x in fresh}
    assert by_id[rows[0]["id"]]["values"]["asn"] == "AS64500"
    assert by_id[rows[1]["id"]]["values"]["asnname"] == "Example"
    assert by_id[rows[2]["id"]]["values"]["date"] == "2026-08-23"


def test_batch_update_broken_row_does_not_kill():
    a = _auth()
    pid, lid = _mk_list(a)
    client.post(f"/api/subnets/providers/{pid}/lists/{lid}/rows", headers=a,
                json={"subnets": ["203.0.113.0/24"]})
    rid = _rows(a)[0]["id"]
    r = client.patch(f"/api/subnets/providers/{pid}/lists/{lid}/rows/batch", headers=a,
                     json={"updates": [
                         {"row_id": "nope", "col": "asn", "value": "X"},
                         {"row_id": rid, "col": "asn", "value": "AS1"},
                         {"row_id": rid, "col": "", "value": "Y"},
                     ]})
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["updated"] == 1 and body["skipped"] == 2 and len(body["errors"]) == 2
    assert _rows(a)[0]["values"]["asn"] == "AS1"


def test_batch_update_limits_and_404():
    a = _auth()
    pid, lid = _mk_list(a)
    updates = [{"row_id": f"r{i}", "col": "asn", "value": "x"} for i in range(501)]
    r = client.patch(f"/api/subnets/providers/{pid}/lists/{lid}/rows/batch",
                     headers=a, json={"updates": updates})
    assert r.status_code == 422  # больше 500
    r = client.patch(f"/api/subnets/providers/{pid}/lists/{lid}/rows/batch",
                     headers=a, json={"updates": []})
    assert r.status_code == 422  # пусто
    r = client.patch("/api/subnets/providers/nope/lists/nope2/rows/batch",
                     headers=a, json={"updates": [{"row_id": "r", "col": "c"}]})
    assert r.status_code == 404


def test_import_json_rows_with_metadata():
    a = _auth()
    pid, lid = _mk_list(a)
    rows = [{"subnet": f"10.{i}.0.0/24", "country": "RU", "asn": f"AS{1000 + i}"}
            for i in range(10)]
    r = client.post(f"/api/subnets/providers/{pid}/lists/{lid}/import-json", headers=a,
                    json={"rows": rows})
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["ok"] is True and body["added"] == 10 and body["skipped"] == 0
    fresh = _rows(a)
    assert len(fresh) == 10
    assert fresh[0]["values"]["country"] == "RU"
    assert fresh[0]["values"]["asn"] == "AS1000"
    # колонки автоматически не создаются
    lst = client.get("/api/subnets", headers=a).json()["providers"][0]["lists"][0]
    assert [c["key"] for c in lst["columns"]] == ["subnet", "ipver", "asn", "asnname", "date", "operators"]


def test_import_json_skips_duplicates_and_bad_rows():
    a = _auth()
    pid, lid = _mk_list(a)
    r = client.post(f"/api/subnets/providers/{pid}/lists/{lid}/import-json", headers=a,
                    json={"rows": [
                        {"subnet": "203.0.113.0/24", "country": "RU"},
                        {"subnet": "203.0.113.0/24", "country": "DE"},  # дубль
                        {"subnet": "мусор"},                            # битая
                        {"subnet": "198.51.100.0/24", "asnname": "Example"},
                    ]})
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["ok"] is True and body["added"] == 2 and body["skipped"] == 2
    assert len(body["errors"]) == 1
    fresh = _rows(a)
    by_subnet = {x["values"]["subnet"]: x for x in fresh}
    assert set(by_subnet) == {"203.0.113.0/24", "198.51.100.0/24"}
    assert by_subnet["198.51.100.0/24"]["values"]["asnname"] == "Example"
    # дубль против уже существующих строк
    r = client.post(f"/api/subnets/providers/{pid}/lists/{lid}/import-json", headers=a,
                    json={"rows": [{"subnet": "203.0.113.0/24"}]})
    assert r.json()["added"] == 0 and r.json()["skipped"] == 1
    # пустой rows — 422, нет списка — 404
    assert client.post(f"/api/subnets/providers/{pid}/lists/{lid}/import-json",
                       headers=a, json={"rows": []}).status_code == 422
    assert client.post("/api/subnets/providers/nope/lists/nope2/import-json",
                       headers=a, json={"rows": [{"subnet": "1.2.3.0/24"}]}).status_code == 404


# ── экспорт xlsx (Excel, группы по провайдеру) ──────────────────
def _xlsx_list(a, rows):
    """Список со строками {subnet, provider?, operator?} → (wb, ws)."""
    pid, lid = _mk_list(a)
    r = client.post(f"/api/subnets/providers/{pid}/lists/{lid}/rows",
                    headers=a, json={"rows": rows})
    assert r.status_code == 201, r.text
    r = client.get("/api/subnets/export", headers=a,
                   params={"provider_id": pid, "list_id": lid, "format": "xlsx"})
    assert r.status_code == 200, r.text
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    return wb, wb.active


def test_export_xlsx_ok_zip_magic():
    a = _auth()
    pid, lid = _mk_list(a)
    client.post(f"/api/subnets/providers/{pid}/lists/{lid}/rows", headers=a,
                json={"subnets": ["203.0.113.0/24"]})
    r = client.get("/api/subnets/export", headers=a,
                   params={"provider_id": pid, "list_id": lid, "format": "xlsx"})
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert "attachment" in r.headers["content-disposition"]
    assert ".xlsx" in r.headers["content-disposition"]
    assert r.content[:4] == b"PK\x03\x04"  # zip-магия xlsx
    # пустой список тоже валиден: только заголовок, без групп
    pid2, lid2 = _mk_list(a)
    r = client.get("/api/subnets/export", headers=a,
                   params={"list_id": lid2, "format": "xlsx"})
    assert r.status_code == 200 and r.content[:4] == b"PK\x03\x04"
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    assert wb.active.max_row == 1


def test_export_xlsx_requires_list_id():
    a = _auth()
    r = client.get("/api/subnets/export", headers=a, params={"format": "xlsx"})
    assert r.status_code == 400
    assert "list_id" in r.json()["detail"]


def test_export_xlsx_groups_by_provider():
    a = _auth()
    rows = [
        {"subnet": "203.0.113.0/24", "provider": "Ростелеком"},
        {"subnet": "198.51.100.0/24", "provider": "МТС"},
        {"subnet": "192.0.2.0/24", "provider": "Билайн"},
        {"subnet": "10.10.0.0/16", "provider": "МТС"},
        {"subnet": "2001:db8::/32"},  # без провайдера — в конец
    ]
    wb, ws = _xlsx_list(a, rows)
    # заголовок: табличные колонки, operators → 5 колонок с лейблами
    head = [ws.cell(1, c).value for c in range(1, 11)]
    assert head == ["subnet", "ipver", "asn", "asnname", "date",
                    "MTS", "Beeline", "МегаФон", "Tele2", "T-Mobile"]
    assert ws["A1"].font.bold
    assert ws.freeze_panes == "A2"
    # Строки сгруппированы по провайдеру: заголовок группы (уровень 0,
    # bold) + данные (уровень 1). Порядок: Билайн → МТС → Ростелеком → «—».
    assert [ws.cell(r, 1).value for r in range(2, 11)] == [
        "Билайн (1)", "192.0.2.0/24",
        "МТС (2)", "198.51.100.0/24", "10.10.0.0/16",
        "Ростелеком (1)", "203.0.113.0/24",
        "— (1)", "2001:db8::/32",
    ]
    # заголовки групп — жирные, уровень 0
    for r in (2, 4, 7, 9):
        assert ws.cell(r, 1).font.bold, f"заголовок группы строка {r}"
        assert ws.row_dimensions[r].outline_level == 0
    # данные групп — уровень 1, hidden=False (сворачивание «+/-»)
    for r in (3, 5, 6, 8, 10):
        dim = ws.row_dimensions[r]
        assert dim.outline_level == 1, f"строка {r} не сгруппирована"
        assert dim.hidden is False
    assert ws.max_row == 10


def test_export_xlsx_values_match_source():
    a = _auth()
    rows = [
        {"subnet": "203.0.113.9/24", "provider": "МТС", "operator": "МТС"},
        {"subnet": "198.51.100.0/24", "provider": "Билайн",
         "operator": "Билайн + Tele2"},
    ]
    _, ws = _xlsx_list(a, rows)
    # первая группа — «Билайн»: заголовок (строка 2), данные (строка 3)
    assert ws["A2"].value == "Билайн (1)"
    assert ws["A3"].value == "198.51.100.0/24"
    assert ws["B3"].value == "IPv4"
    # операторы: 1/0, как в CSV
    assert [ws.cell(3, c).value for c in range(6, 11)] == [0, 1, 0, 1, 0]
    assert [ws.cell(5, c).value for c in range(6, 11)] == [1, 0, 0, 0, 0]
    # группы: строка 3 (Билайн) и строка 5 (МТС) — данные уровня 1
    assert ws.row_dimensions[3].outline_level == 1
    assert ws.row_dimensions[5].outline_level == 1


# ── тип ASN (эвристика isp/hosting/business) ──────────────────
def test_asn_type_heuristic():
    t = store._asn_type
    # hosting — приоритетнее isp
    assert t("Selectel", "SELECTEL-NET", "SELECTEL-NET", "") == "hosting"
    assert t("", "Timeweb Cloud", "", "") == "hosting"
    assert t("ООО «Облако»", "", "", "") == "hosting"
    assert t("", "", "AEZA-Data-Center", "") == "hosting"  # дефис == пробел
    assert t("Hetzner Online GmbH", "", "", "") == "hosting"
    assert t("", "DigitalOcean", "", "") == "hosting"
    assert t("Amazon Web Services", "", "", "") == "business"  # «amazon» не в списке
    assert t("", "AWS-AMAZON", "", "") == "hosting"  # ключевое слово aws
    assert t("", "", "", "RUVDS") == "hosting"  # по провайдеру
    assert t("RU-SERVER", "RU-SERVER", "", "") == "hosting"  # "server" в строке
    # isp
    assert t("Ростелеком", "", "", "") == "isp"
    assert t("", "", "", "МТС") == "isp"
    assert t("PJSC MegaFon", "", "", "") == "isp"
    assert t("", "Beeline", "", "") == "isp"
    assert t("ОАО «Транстелеком»", "", "", "") == "isp"
    assert t("Some Telecom LLC", "", "", "") == "isp"
    assert t("", "", "INTERNET-NET", "") == "isp"
    assert t("Wireline Networks", "", "", "") == "isp"
    # business — всё остальное с данными
    assert t("Газпром", "GAZPROM-TRADE", "", "") == "business"
    assert t("", "CorpSystems", "CORP-1", "") == "business"
    # пусто — данных нет
    assert t("", "", "", "") == ""
    # регистр не важен
    assert t("HOSTING", "", "", "") == "hosting"
    assert t("", "hEtZnEr", "", "") == "hosting"


def test_enrich_types_fills_column_without_touching_provider():
    a = _auth()
    pid, lid = _mk_list(a)
    client.post(f"/api/subnets/providers/{pid}/lists/{lid}/rows", headers=a,
                json={"rows": [
                    {"subnet": "203.0.113.0/24", "org": "Selectel",
                     "asnname": "SELECTEL-NET", "provider": "МТС"},
                    {"subnet": "198.51.100.0/24", "org": "Ростелеком"},
                    {"subnet": "192.0.2.0/24", "org": "Газпром"},
                    {"subnet": "10.0.0.0/8"},  # данных нет — пропуск
                ]})
    r = client.post(f"/api/subnets/providers/{pid}/lists/{lid}/enrich-types",
                    headers=a, json={})
    assert r.status_code == 200
    assert r.json() == {"updated": 3, "of": 4}
    data = client.get("/api/subnets", headers=a).json()
    lst = data["providers"][0]["lists"][0]
    # колонка asn_type создана (реальная колонка → уходит в экспорт)
    assert any(c["key"] == "asn_type" and c["title"] == "Тип ASN"
               for c in lst["columns"])
    by_subnet = {x["values"]["subnet"]: x["values"] for x in lst["rows"]}
    assert by_subnet["203.0.113.0/24"]["asn_type"] == "hosting"
    assert by_subnet["198.51.100.0/24"]["asn_type"] == "isp"
    assert by_subnet["192.0.2.0/24"]["asn_type"] == "business"
    assert "asn_type" not in by_subnet["10.0.0.0/8"]
    # provider НЕ тронут
    assert by_subnet["203.0.113.0/24"]["provider"] == "МТС"
    # повторный вызов не дублирует колонку, asn_type в CSV-экспорте
    r = client.post(f"/api/subnets/providers/{pid}/lists/{lid}/enrich-types",
                    headers=a, json={})
    assert r.json() == {"updated": 0, "of": 4}
    cols = [c["key"] for c in lst["columns"]]
    assert cols.count("asn_type") == 1
    r = client.get("/api/subnets/export", headers=a,
                   params={"provider_id": pid, "list_id": lid, "format": "csv"})
    assert r.status_code == 200
    assert "Тип ASN" in r.text.splitlines()[0]


def test_enrich_types_row_ids_and_404():
    a = _auth()
    pid, lid = _mk_list(a)
    client.post(f"/api/subnets/providers/{pid}/lists/{lid}/rows", headers=a,
                json={"rows": [
                    {"subnet": "203.0.113.0/24", "org": "Selectel"},
                    {"subnet": "198.51.100.0/24", "org": "Ростелеком"},
                ]})
    rows = _rows(a)
    rid = next(x["id"] for x in rows if x["values"]["subnet"] == "203.0.113.0/24")
    # только выбранная строка
    r = client.post(f"/api/subnets/providers/{pid}/lists/{lid}/enrich-types",
                    headers=a, json={"row_ids": [rid]})
    assert r.status_code == 200 and r.json() == {"updated": 1, "of": 1}
    fresh = _rows(a)
    by_subnet = {x["values"]["subnet"]: x["values"] for x in fresh}
    assert by_subnet["203.0.113.0/24"]["asn_type"] == "hosting"
    assert "asn_type" not in by_subnet["198.51.100.0/24"]
    # несуществующие строки/список
    assert client.post(f"/api/subnets/providers/{pid}/lists/{lid}/enrich-types",
                       headers=a, json={"row_ids": ["nope"]}).status_code == 404
    assert client.post("/api/subnets/providers/nope/lists/nope2/enrich-types",
                       headers=a, json={}).status_code == 404


# ── иконки провайдеров/списков ────────────────────────────────
_PNG = b"\x89PNG\r\n\x1a\n" + b"0" * 64


def test_icon_upload_and_serve_provider_and_list():
    a = _auth()
    pid, lid = _mk_list(a)
    # до загрузки — иконки нет, GET отдаёт 404
    assert client.get(f"/api/subnets/provider-icon/{pid}", headers=a).status_code == 404
    assert client.get(f"/api/subnets/list-icon/{pid}/{lid}", headers=a).status_code == 404
    assert client.get("/api/subnets/provider-icon/nope", headers=a).status_code == 404

    # загрузка иконки провайдера (multipart png)
    r = client.post(f"/api/subnets/providers/{pid}/icon", headers=a,
                    files={"file": ("icon.png", _PNG, "image/png")})
    assert r.status_code == 200 and r.json()["ok"] is True
    data = client.get("/api/subnets", headers=a).json()
    prov = data["providers"][0]
    assert prov["icon"] == f"{pid}.png"
    # GET отдаёт файл с правильным content-type
    r = client.get(f"/api/subnets/provider-icon/{pid}", headers=a)
    assert r.status_code == 200
    assert r.content == _PNG
    assert r.headers["content-type"].startswith("image/png")

    # иконка списка — отдельно
    r = client.post(f"/api/subnets/providers/{pid}/lists/{lid}/icon", headers=a,
                    files={"file": ("list.svg", b"<svg/>", "image/svg+xml")})
    assert r.status_code == 200
    data = client.get("/api/subnets", headers=a).json()
    lst = data["providers"][0]["lists"][0]
    assert lst["icon"] == f"list_{lid}.svg"
    r = client.get(f"/api/subnets/list-icon/{pid}/{lid}", headers=a)
    assert r.status_code == 200 and r.content == b"<svg/>"
    assert "image/svg" in r.headers["content-type"]

    # перезапись с другим расширением: старый файл заменяется
    r = client.post(f"/api/subnets/providers/{pid}/icon", headers=a,
                    files={"file": ("icon.webp", b"WEBP", "image/webp")})
    assert r.status_code == 200
    prov = client.get("/api/subnets", headers=a).json()["providers"][0]
    assert prov["icon"] == f"{pid}.webp"
    assert client.get(f"/api/subnets/provider-icon/{pid}", headers=a).content == b"WEBP"


def test_icon_upload_validation():
    a = _auth()
    pid, lid = _mk_list(a)
    # не тот формат
    r = client.post(f"/api/subnets/providers/{pid}/icon", headers=a,
                    files={"file": ("icon.jpg", b"jpeg", "image/jpeg")})
    assert r.status_code == 400 and "PNG" in r.json()["detail"]
    # пустой файл
    r = client.post(f"/api/subnets/providers/{pid}/icon", headers=a,
                    files={"file": ("icon.png", b"", "image/png")})
    assert r.status_code == 400
    # больше 256 КБ
    big = b"x" * (256 * 1024 + 1)
    r = client.post(f"/api/subnets/providers/{pid}/lists/{lid}/icon", headers=a,
                    files={"file": ("big.png", big, "image/png")})
    assert r.status_code == 400 and "256" in r.json()["detail"]
    # несуществующий провайдер/список — 404
    r = client.post("/api/subnets/providers/nope/icon", headers=a,
                    files={"file": ("icon.png", _PNG, "image/png")})
    assert r.status_code == 404
    # файл не сохранился после битой загрузки
    assert client.get(f"/api/subnets/provider-icon/{pid}", headers=a).status_code == 404


# ── справочник ASN (per-account, синхронизация provider в строках) ──
def _mk_rows_with_asn(a, pid, lid, asn1="AS12345", asn2="AS999"):
    r = client.post(f"/api/subnets/providers/{pid}/lists/{lid}/rows", headers=a,
                    json={"rows": [
                        {"subnet": "203.0.113.0/24", "asn": asn1},
                        {"subnet": "198.51.100.0/24", "asn": asn2},
                    ]})
    assert r.status_code == 201, r.text
    return {x["values"]["subnet"]: x["values"] for x in _rows(a)}


def test_asns_upsert_normalizes_without_as_prefix():
    a = _auth()
    # «12345» без префикса → «AS12345»; регистр/пробелы не важны
    r = client.post("/api/subnets/asns", headers=a,
                    json={"asn": "12345", "name": "Яндекс", "note": "основной"})
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["ok"] is True
    assert body["asn"]["asn"] == "AS12345"
    assert body["asn"]["name"] == "Яндекс"
    assert body["asn"]["note"] == "основной"
    assert body["updated_rows"] == 0
    # GET отдаёт нормализованную запись
    asns = client.get("/api/subnets/asns", headers=a).json()["asns"]
    assert any(x["asn"] == "AS12345" and x["name"] == "Яндекс" for x in asns)
    # «as12345» и «AS 12345» — тот же ключ (upsert, не дубль)
    client.post("/api/subnets/asns", headers=a,
                json={"asn": "as 12345", "name": "Яндекс", "note": ""})
    asns = client.get("/api/subnets/asns", headers=a).json()["asns"]
    assert sum(1 for x in asns if x["asn"] == "AS12345") == 1
    # мусор — 422
    r = client.post("/api/subnets/asns", headers=a, json={"asn": "abc"})
    assert r.status_code == 422


def test_asns_upsert_syncs_provider_in_all_lists():
    """Upsert справочника перезаписывает values.provider у ВСЕХ строк с
    values.asn == «AS12345» (во всех провайдерах/списках); asnname НЕ
    трогается (это поле больше не связано со справочником); чужие ASN не
    трогаются; ответ содержит updated_rows."""
    a = _auth()
    pid, lid = _mk_list(a)
    _mk_rows_with_asn(a, pid, lid)
    # вторая пара провайдер/список — синхронизация идёт по всем спискам
    p2 = client.post("/api/subnets/providers", headers=a, json={"name": "Билайн"}).json()
    l2 = client.post(f"/api/subnets/providers/{p2['id']}/lists", headers=a,
                     json={"name": "Второй"}).json()
    _mk_rows_with_asn(a, p2["id"], l2["id"])

    r = client.post("/api/subnets/asns", headers=a,
                    json={"asn": "12345", "name": "Яндекс"})
    assert r.status_code == 200
    assert r.json()["updated_rows"] == 2  # по одной строке в каждом списке

    data = client.get("/api/subnets", headers=a).json()
    for p in data["providers"]:
        for l in p["lists"]:
            by_sub = {x["values"]["subnet"]: x["values"] for x in l["rows"]}
            assert by_sub["203.0.113.0/24"]["provider"] == "Яндекс"
            assert "asnname" not in by_sub["203.0.113.0/24"]  # asnname не тронут
            assert "provider" not in by_sub["198.51.100.0/24"]  # AS999 не тронут


def test_asns_reupsert_updates_name_and_rows():
    a = _auth()
    pid, lid = _mk_list(a)
    _mk_rows_with_asn(a, pid, lid)
    client.post("/api/subnets/asns", headers=a, json={"asn": "12345", "name": "Яндекс"})
    # повторный upsert с другим названием → запись и provider в строках обновлены
    r = client.post("/api/subnets/asns", headers=a,
                    json={"asn": "AS12345", "name": "Яндекс Облако"})
    assert r.status_code == 200
    assert r.json()["updated_rows"] == 1
    asns = client.get("/api/subnets/asns", headers=a).json()["asns"]
    rec = next(x for x in asns if x["asn"] == "AS12345")
    assert rec["name"] == "Яндекс Облако"
    fresh = {x["values"]["subnet"]: x["values"] for x in _rows(a)}
    assert fresh["203.0.113.0/24"]["provider"] == "Яндекс Облако"
    # пустое name при апдейте не затирает название
    client.post("/api/subnets/asns", headers=a, json={"asn": "12345", "name": ""})
    rec = next(x for x in client.get("/api/subnets/asns", headers=a).json()["asns"]
               if x["asn"] == "AS12345")
    assert rec["name"] == "Яндекс Облако"


def test_asns_delete_removes_record_keeps_row_values():
    """DELETE /asns/{asn}: запись уходит из справочника, значения в строках
    подсетей НЕ трогаются (provider остаётся последним названием)."""
    a = _auth()
    pid, lid = _mk_list(a)
    _mk_rows_with_asn(a, pid, lid)
    client.post("/api/subnets/asns", headers=a, json={"asn": "12345", "name": "Яндекс"})

    r = client.delete("/api/subnets/asns/AS12345", headers=a)
    assert r.status_code == 200 and r.json()["ok"] is True
    asns = client.get("/api/subnets/asns", headers=a).json()["asns"]
    assert all(x["asn"] != "AS12345" for x in asns)
    # строки не тронуты: provider остался
    fresh = {x["values"]["subnet"]: x["values"] for x in _rows(a)}
    assert fresh["203.0.113.0/24"]["provider"] == "Яндекс"
    # удаление без префикса тоже работает; повторное удаление — идемпотентно
    assert client.delete("/api/subnets/asns/12345", headers=a).status_code == 200
    assert client.delete("/api/subnets/asns/12345", headers=a).status_code == 200
    # мусор в пути — 422
    assert client.delete("/api/subnets/asns/abc", headers=a).status_code == 422


def test_asns_isolated_per_account():
    """Справочник per-account: записи одного аккаунта не видны другому."""
    a1, a2 = _auth(), _auth()
    client.post("/api/subnets/asns", headers=a1, json={"asn": "12345", "name": "Яндекс"})
    assert client.get("/api/subnets/asns", headers=a2).json()["asns"] == []
    # синхронизация тоже по аккаунту: строка a2 с тем же ASN не тронута
    pid, lid = _mk_list(a2)
    _mk_rows_with_asn(a2, pid, lid)
    client.post("/api/subnets/asns", headers=a2, json={"asn": "12345", "name": "Другой"})
    fresh = {x["values"]["subnet"]: x["values"] for x in _rows(a2)}
    assert fresh["203.0.113.0/24"]["provider"] == "Другой"
    assert client.get("/api/subnets/asns", headers=a1).json()["asns"][0]["name"] == "Яндекс"


# ── синхронизация справочника из списков подсетей (POST /asns/sync) ──
def _mk_rows_with_provider(a, pid, lid):
    """Строки для sync: валидная пара (asn+provider), asn без названия,
    мусорный asn, строка без asn вовсе."""
    r = client.post(f"/api/subnets/providers/{pid}/lists/{lid}/rows", headers=a,
                    json={"rows": [
                        {"subnet": "203.0.113.0/24", "asn": "AS12345", "provider": "Яндекс"},
                        {"subnet": "198.51.100.0/24", "asn": "AS3261"},
                        {"subnet": "192.0.2.0/24", "asn": "мусор", "provider": "Мусор"},
                        {"subnet": "10.0.0.0/8"},
                    ]})
    assert r.status_code == 201, r.text


def test_asns_sync_collects_from_lists_adds_missing():
    """POST /asns/sync: ВСЕ уникальные values.asn из ВСЕХ списков попадают в
    справочник — даже без provider (запись с пустым name); мусорный asn и
    строка без asn пропускаются; повторный sync не дублирует записи
    (added=0); строки подсетей не меняются."""
    a = _auth()
    pid, lid = _mk_list(a)
    _mk_rows_with_provider(a, pid, lid)
    # вторая пара провайдер/список — sync идёт по всем спискам
    p2 = client.post("/api/subnets/providers", headers=a, json={"name": "Билайн"}).json()
    l2 = client.post(f"/api/subnets/providers/{p2['id']}/lists", headers=a,
                     json={"name": "Второй"}).json()
    client.post(f"/api/subnets/providers/{p2['id']}/lists/{l2['id']}/rows", headers=a,
                json={"rows": [{"subnet": "2001:db8::/32", "asn": "as999",
                                "provider": "Другой"}]})  # нижний регистр тоже норм

    r = client.post("/api/subnets/asns/sync", headers=a)
    assert r.status_code == 200
    body = r.json()
    assert body["ok"] is True
    assert body["added"] == 3 and body["filled"] == 0 and body["total"] == 3
    by_asn = {x["asn"]: x["name"] for x in client.get("/api/subnets/asns",
                                                      headers=a).json()["asns"]}
    assert by_asn == {"AS12345": "Яндекс", "AS3261": "", "AS999": "Другой"}
    # ASN без названия добавлен записью с пустым name
    rec = next(x for x in client.get("/api/subnets/asns", headers=a).json()["asns"]
               if x["asn"] == "AS3261")
    assert rec["name"] == "" and rec["note"] == ""
    # строки подсетей не тронуты: provider в них остался как был
    fresh = {x["values"]["subnet"]: x["values"] for x in _rows(a)}
    assert fresh["203.0.113.0/24"]["provider"] == "Яндекс"

    # повторный sync — дубликатов нет (и пустое name не перезаполняется)
    r2 = client.post("/api/subnets/asns/sync", headers=a)
    assert r2.json() == {"ok": True, "added": 0, "filled": 0, "total": 3}


def test_asns_sync_keeps_dictionary_name_over_rows():
    """Существующая запись с непустым name не перезаписывается из строк
    (справочник авторитетнее); отсутствующие — добавляются."""
    a = _auth()
    pid, lid = _mk_list(a)
    client.post(f"/api/subnets/providers/{pid}/lists/{lid}/rows", headers=a,
                json={"rows": [
                    {"subnet": "203.0.113.0/24", "asn": "AS3261", "provider": "Новое"},
                    {"subnet": "198.51.100.0/24", "asn": "AS12345", "provider": "Яндекс"},
                ]})
    # в справочнике AS3261 уже есть с другим названием
    client.post("/api/subnets/asns", headers=a, json={"asn": "AS3261", "name": "Старое"})

    r = client.post("/api/subnets/asns/sync", headers=a)
    assert r.status_code == 200
    assert r.json()["added"] == 1 and r.json()["filled"] == 0
    asns = client.get("/api/subnets/asns", headers=a).json()["asns"]
    rec = next(x for x in asns if x["asn"] == "AS3261")
    assert rec["name"] == "Старое"  # НЕ перезаписано из файла
    assert any(x["asn"] == "AS12345" and x["name"] == "Яндекс" for x in asns)


def test_asns_sync_fills_empty_name_in_dictionary():
    """У существующей записи с пустым name название заполняется из строк;
    примечание при этом сохраняется."""
    a = _auth()
    pid, lid = _mk_list(a)
    client.post(f"/api/subnets/providers/{pid}/lists/{lid}/rows", headers=a,
                json={"rows": [{"subnet": "203.0.113.0/24", "asn": "12345",
                                "provider": "Яндекс"}]})
    # запись есть, но без названия (и с примечанием)
    client.post("/api/subnets/asns", headers=a,
                json={"asn": "12345", "name": "", "note": "важно"})

    r = client.post("/api/subnets/asns/sync", headers=a)
    assert r.status_code == 200
    assert r.json()["added"] == 0 and r.json()["filled"] == 1
    rec = next(x for x in client.get("/api/subnets/asns", headers=a).json()["asns"]
               if x["asn"] == "AS12345")
    assert rec["name"] == "Яндекс"
    assert rec["note"] == "важно"  # примечание не затёрто


def test_asns_sync_respects_max(monkeypatch):
    """Лимит MAX_ASNS: когда справочник упирается в потолок, sync добавляет
    сколько влезает и не падает."""
    import app.services.asn_store as asn_store
    a = _auth()
    pid, lid = _mk_list(a)
    client.post(f"/api/subnets/providers/{pid}/lists/{lid}/rows", headers=a,
                json={"rows": [
                    {"subnet": "203.0.113.0/24", "asn": "AS1", "provider": "Один"},
                    {"subnet": "198.51.100.0/24", "asn": "AS2", "provider": "Два"},
                    {"subnet": "192.0.2.0/24", "asn": "AS3", "provider": "Три"},
                ]})
    monkeypatch.setattr(asn_store, "MAX_ASNS", 2)
    r = client.post("/api/subnets/asns/sync", headers=a)
    assert r.status_code == 200
    body = r.json()
    assert body["added"] == 2 and body["total"] == 2


def test_asns_upsert_netname_saved_and_not_cleared():
    """Upsert с netname: запись получает поле netname; апдейт с пустым
    netname НЕ затирает текущее; строки подсетей получают provider И netname
    (apply_asn_meta), updated_rows > 0."""
    a = _auth()
    pid, lid = _mk_list(a)
    _mk_rows_with_asn(a, pid, lid)
    r = client.post("/api/subnets/asns", headers=a,
                    json={"asn": "12345", "name": "Яндекс", "netname": "RU-YANDEX"})
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["asn"]["netname"] == "RU-YANDEX"
    assert body["updated_rows"] == 1
    # повторный upsert с пустым netname — текущее значение сохраняется
    client.post("/api/subnets/asns", headers=a,
                json={"asn": "12345", "name": "", "netname": ""})
    rec = next(x for x in client.get("/api/subnets/asns", headers=a).json()["asns"]
               if x["asn"] == "AS12345")
    assert rec["netname"] == "RU-YANDEX"
    assert rec["name"] == "Яндекс"
    # из справочника в строки перенесены и provider, и netname
    fresh = {x["values"]["subnet"]: x["values"] for x in _rows(a)}
    assert fresh["203.0.113.0/24"]["provider"] == "Яндекс"
    assert fresh["203.0.113.0/24"]["netname"] == "RU-YANDEX"
    # чужой ASN не тронут
    assert "netname" not in fresh["198.51.100.0/24"]


def test_asns_apply_applies_whole_dictionary_idempotent():
    """POST /asns/apply: ВЕСЬ справочник (name и netname) переносится в строки
    подсетей; повторный apply — идемпотентен (updated_rows тот же)."""
    a = _auth()
    pid, lid = _mk_list(a)
    _mk_rows_with_asn(a, pid, lid)
    # одна запись с name+netname, другая — только с netname
    client.post("/api/subnets/asns", headers=a,
                json={"asn": "12345", "name": "Яндекс", "netname": "RU-YANDEX"})
    client.post("/api/subnets/asns", headers=a,
                json={"asn": "999", "name": "", "netname": "RU-OTHER"})

    r = client.post("/api/subnets/asns/apply", headers=a)
    assert r.status_code == 200
    body = r.json()
    assert body["ok"] is True and body["updated_rows"] == 2
    fresh = {x["values"]["subnet"]: x["values"] for x in _rows(a)}
    assert fresh["203.0.113.0/24"]["provider"] == "Яндекс"
    assert fresh["203.0.113.0/24"]["netname"] == "RU-YANDEX"
    assert fresh["198.51.100.0/24"]["netname"] == "RU-OTHER"
    # у записи без name provider в строках не появляется
    assert "provider" not in fresh["198.51.100.0/24"]

    # повторный apply — снова перезаписывает те же строки (идемпотентно)
    r2 = client.post("/api/subnets/asns/apply", headers=a)
    assert r2.json()["updated_rows"] == 2


def test_asns_apply_skips_rows_not_in_dictionary():
    """apply не трогает строки, чей ASN отсутствует в справочнике."""
    a = _auth()
    pid, lid = _mk_list(a)
    _mk_rows_with_asn(a, pid, lid)  # AS12345 и AS999 в строках
    client.post("/api/subnets/asns", headers=a,
                json={"asn": "12345", "name": "Яндекс", "netname": "RU-YANDEX"})
    r = client.post("/api/subnets/asns/apply", headers=a)
    assert r.json()["updated_rows"] == 1
    fresh = {x["values"]["subnet"]: x["values"] for x in _rows(a)}
    assert fresh["203.0.113.0/24"]["provider"] == "Яндекс"
    # AS999 в справочнике нет — строка не тронута
    assert "netname" not in fresh["198.51.100.0/24"]
    assert "provider" not in fresh["198.51.100.0/24"]


def test_asns_sync_collects_netname():
    """POST /asns/sync собирает values.netname из строк в справочник; строки
    подсетей при этом не меняются; повторный sync — без дубликатов."""
    a = _auth()
    pid, lid = _mk_list(a)
    client.post(f"/api/subnets/providers/{pid}/lists/{lid}/rows", headers=a,
                json={"rows": [
                    {"subnet": "203.0.113.0/24", "asn": "AS12345",
                     "provider": "Яндекс", "netname": "RU-YANDEX"},
                    {"subnet": "198.51.100.0/24", "asn": "AS3261"},
                ]})
    r = client.post("/api/subnets/asns/sync", headers=a)
    assert r.status_code == 200
    assert r.json()["added"] == 2
    by_asn = {x["asn"]: x for x in
              client.get("/api/subnets/asns", headers=a).json()["asns"]}
    assert by_asn["AS12345"]["name"] == "Яндекс"
    assert by_asn["AS12345"]["netname"] == "RU-YANDEX"
    assert by_asn["AS3261"]["netname"] == ""  # в строке netname не было
    # строки подсетей не тронуты
    fresh = {x["values"]["subnet"]: x["values"] for x in _rows(a)}
    assert fresh["203.0.113.0/24"]["netname"] == "RU-YANDEX"
    # повторный sync — дубликатов нет, netname не перезаписывается
    r2 = client.post("/api/subnets/asns/sync", headers=a)
    assert r2.json() == {"ok": True, "added": 0, "filled": 0, "total": 2}


# ── иконки записей ASN (у ASN, а не у файлов/провайдеров) ─────
def test_asn_icon_upload_serve_and_delete():
    """POST /asns/{asn}/icon кладёт icon в запись; GET /asns/{asn}/icon отдаёт
    файл; GET /asns возвращает icon; DELETE /asns/{asn} удаляет и иконку."""
    a = _auth()
    assert client.post("/api/subnets/asns", headers=a,
                       json={"asn": "12345", "name": "Яндекс"}).status_code == 200
    # новая запись получает пустое поле icon; GET до загрузки — 404
    rec = next(x for x in client.get("/api/subnets/asns", headers=a).json()["asns"]
               if x["asn"] == "AS12345")
    assert rec["icon"] == ""
    assert client.get("/api/subnets/asns/AS12345/icon", headers=a).status_code == 404

    # загрузка иконки (multipart png) → поле icon = asn_AS12345.png
    r = client.post("/api/subnets/asns/AS12345/icon", headers=a,
                    files={"file": ("icon.png", _PNG, "image/png")})
    assert r.status_code == 200 and r.json()["ok"] is True
    rec = next(x for x in client.get("/api/subnets/asns", headers=a).json()["asns"]
               if x["asn"] == "AS12345")
    assert rec["icon"] == "asn_AS12345.png"
    # GET отдаёт файл с правильным content-type
    r = client.get("/api/subnets/asns/AS12345/icon", headers=a)
    assert r.status_code == 200 and r.content == _PNG
    assert r.headers["content-type"].startswith("image/png")

    # перезапись с другим расширением: старый файл заменяется
    r = client.post("/api/subnets/asns/AS12345/icon", headers=a,
                    files={"file": ("icon.svg", b"<svg/>", "image/svg+xml")})
    assert r.status_code == 200
    rec = next(x for x in client.get("/api/subnets/asns", headers=a).json()["asns"]
               if x["asn"] == "AS12345")
    assert rec["icon"] == "asn_AS12345.svg"
    r = client.get("/api/subnets/asns/AS12345/icon", headers=a)
    assert r.status_code == 200 and r.content == b"<svg/>"
    assert "image/svg" in r.headers["content-type"]

    # удаление записи удаляет и иконку: запись ушла, GET icon — 404
    assert client.delete("/api/subnets/asns/AS12345", headers=a).status_code == 200
    assert client.get("/api/subnets/asns/AS12345/icon", headers=a).status_code == 404
    asns = client.get("/api/subnets/asns", headers=a).json()["asns"]
    assert all(x["asn"] != "AS12345" for x in asns)


def test_asn_icon_upload_validation():
    a = _auth()
    client.post("/api/subnets/asns", headers=a, json={"asn": "12345", "name": "Яндекс"})
    # не тот формат
    r = client.post("/api/subnets/asns/AS12345/icon", headers=a,
                    files={"file": ("icon.jpg", b"jpeg", "image/jpeg")})
    assert r.status_code == 400 and "PNG" in r.json()["detail"]
    # пустой файл
    r = client.post("/api/subnets/asns/AS12345/icon", headers=a,
                    files={"file": ("icon.png", b"", "image/png")})
    assert r.status_code == 400
    # больше 256 КБ
    big = b"x" * (256 * 1024 + 1)
    r = client.post("/api/subnets/asns/AS12345/icon", headers=a,
                    files={"file": ("big.png", big, "image/png")})
    assert r.status_code == 400 and "256" in r.json()["detail"]
    # записи нет — 404; мусор в пути — 422
    assert client.post("/api/subnets/asns/AS99999/icon", headers=a,
                       files={"file": ("icon.png", _PNG, "image/png")}).status_code == 404
    assert client.post("/api/subnets/asns/abc/icon", headers=a,
                       files={"file": ("icon.png", _PNG, "image/png")}).status_code == 422
    # файл не сохранился после битой загрузки
    assert client.get("/api/subnets/asns/AS12345/icon", headers=a).status_code == 404


# ── country/asn_type в справочнике + авто-синхронизация строки → справочник ──
def test_asns_upsert_country_asn_type_saved_and_not_cleared():
    """Upsert с country/asn_type: запись получает поля; апдейт с пустыми
    country/asn_type НЕ затирает текущие; строки подсетей получают country
    и asn_type через apply_asn_meta."""
    a = _auth()
    pid, lid = _mk_list(a)
    _mk_rows_with_asn(a, pid, lid)
    r = client.post("/api/subnets/asns", headers=a,
                    json={"asn": "12345", "name": "Яндекс",
                          "country": "RU", "asn_type": "hosting"})
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["asn"]["country"] == "RU"
    assert body["asn"]["asn_type"] == "hosting"
    assert body["updated_rows"] == 1
    # повторный upsert с пустыми country/asn_type — значения сохраняются
    client.post("/api/subnets/asns", headers=a,
                json={"asn": "12345", "name": "", "country": "", "asn_type": ""})
    rec = next(x for x in client.get("/api/subnets/asns", headers=a).json()["asns"]
               if x["asn"] == "AS12345")
    assert rec["country"] == "RU"
    assert rec["asn_type"] == "hosting"
    # из справочника в строки перенесены и country, и asn_type
    fresh = {x["values"]["subnet"]: x["values"] for x in _rows(a)}
    assert fresh["203.0.113.0/24"]["country"] == "RU"
    assert fresh["203.0.113.0/24"]["asn_type"] == "hosting"
    # чужой ASN не тронут
    assert "country" not in fresh["198.51.100.0/24"]


def test_asns_sync_collects_country_and_asn_type():
    """POST /asns/sync собирает values.country и values.asn_type из строк в
    справочник; существующие непустые значения не перезаписываются."""
    a = _auth()
    pid, lid = _mk_list(a)
    client.post(f"/api/subnets/providers/{pid}/lists/{lid}/rows", headers=a,
                json={"rows": [
                    {"subnet": "203.0.113.0/24", "asn": "AS12345",
                     "provider": "Яндекс", "country": "RU", "asn_type": "hosting"},
                    {"subnet": "198.51.100.0/24", "asn": "AS3261",
                     "country": "DE", "asn_type": "isp"},
                ]})
    r = client.post("/api/subnets/asns/sync", headers=a)
    assert r.status_code == 200
    assert r.json()["added"] == 2
    by_asn = {x["asn"]: x for x in
              client.get("/api/subnets/asns", headers=a).json()["asns"]}
    assert by_asn["AS12345"]["country"] == "RU"
    assert by_asn["AS12345"]["asn_type"] == "hosting"
    assert by_asn["AS3261"]["country"] == "DE"
    assert by_asn["AS3261"]["asn_type"] == "isp"
    # существующая запись с непустыми полями не перезаписывается из строк
    client.post("/api/subnets/asns", headers=a,
                json={"asn": "AS12345", "name": "Яндекс",
                      "country": "KZ", "asn_type": "business"})
    r = client.post("/api/subnets/asns/sync", headers=a)
    assert r.json()["added"] == 0 and r.json()["filled"] == 0
    rec = next(x for x in client.get("/api/subnets/asns", headers=a).json()["asns"]
               if x["asn"] == "AS12345")
    assert rec["country"] == "KZ"  # не перезаписано из строк
    assert rec["asn_type"] == "business"


def test_asns_apply_transfers_country_and_asn_type():
    """POST /asns/apply переносит country/asn_type из справочника в строки
    подсетей (как name/netname); повторный apply идемпотентен."""
    a = _auth()
    pid, lid = _mk_list(a)
    _mk_rows_with_asn(a, pid, lid)
    client.post("/api/subnets/asns", headers=a,
                json={"asn": "12345", "name": "Яндекс",
                      "country": "RU", "asn_type": "hosting"})
    r = client.post("/api/subnets/asns/apply", headers=a)
    assert r.status_code == 200
    assert r.json()["updated_rows"] == 1
    fresh = {x["values"]["subnet"]: x["values"] for x in _rows(a)}
    assert fresh["203.0.113.0/24"]["country"] == "RU"
    assert fresh["203.0.113.0/24"]["asn_type"] == "hosting"
    # AS999 в справочнике нет — строка не тронута
    assert "country" not in fresh["198.51.100.0/24"]
    # повторный apply — снова те же строки (идемпотентно)
    r2 = client.post("/api/subnets/asns/apply", headers=a)
    assert r2.json()["updated_rows"] == 1


def test_batch_cell_edit_syncs_row_to_dictionary():
    """Правка ячейки provider/netname/country/asn_type строки с ASN из
    справочника АВТОМАТИЧЕСКИ обновляет запись справочника (строки →
    справочник, без apply обратно)."""
    a = _auth()
    pid, lid = _mk_list(a)
    client.post(f"/api/subnets/providers/{pid}/lists/{lid}/rows", headers=a,
                json={"rows": [
                    {"subnet": "203.0.113.0/24", "asn": "AS12345"},
                    {"subnet": "198.51.100.0/24", "asn": "AS999"},
                ]})
    client.post("/api/subnets/asns", headers=a,
                json={"asn": "12345", "name": "Яндекс"})
    rows = _rows(a)
    rid1 = next(x["id"] for x in rows if x["values"]["subnet"] == "203.0.113.0/24")
    r = client.patch(f"/api/subnets/providers/{pid}/lists/{lid}/rows/batch",
                     headers=a, json={"updates": [
                         {"row_id": rid1, "col": "provider", "value": "Яндекс Облако"},
                         {"row_id": rid1, "col": "netname", "value": "RU-YANDEX"},
                         {"row_id": rid1, "col": "country", "value": "RU"},
                         {"row_id": rid1, "col": "asn_type", "value": "hosting"},
                     ]})
    assert r.status_code == 200, r.text
    assert r.json()["updated"] == 4
    asns = client.get("/api/subnets/asns", headers=a).json()["asns"]
    assert len(asns) == 1  # AS999 в справочник не добавлен
    rec = asns[0]
    assert rec["name"] == "Яндекс Облако"
    assert rec["netname"] == "RU-YANDEX"
    assert rec["country"] == "RU"
    assert rec["asn_type"] == "hosting"
    # строка тоже получила новое значение (ячейка изменена как обычно)
    fresh = {x["values"]["subnet"]: x["values"] for x in _rows(a)}
    assert fresh["203.0.113.0/24"]["provider"] == "Яндекс Облако"
    assert fresh["203.0.113.0/24"]["country"] == "RU"


def test_batch_cell_edit_asnname_does_not_touch_dictionary():
    """Правка ячейки asnname у строки с ASN из справочника НЕ трогает запись
    справочника (asnname больше не синхронизируется с name)."""
    a = _auth()
    pid, lid = _mk_list(a)
    client.post(f"/api/subnets/providers/{pid}/lists/{lid}/rows", headers=a,
                json={"rows": [{"subnet": "203.0.113.0/24", "asn": "AS12345"}]})
    client.post("/api/subnets/asns", headers=a,
                json={"asn": "12345", "name": "Яндекс"})
    rid = _rows(a)[0]["id"]
    r = client.patch(f"/api/subnets/providers/{pid}/lists/{lid}/rows/batch",
                     headers=a, json={"updates": [
                         {"row_id": rid, "col": "asnname", "value": "Яндекс Облако"},
                     ]})
    assert r.status_code == 200, r.text
    assert r.json()["updated"] == 1
    # запись справочника НЕ изменена (name остался «Яндекс»)
    asns = client.get("/api/subnets/asns", headers=a).json()["asns"]
    assert len(asns) == 1
    assert asns[0]["name"] == "Яндекс"
    # ячейка строки при этом изменилась как обычно
    fresh = _rows(a)[0]["values"]
    assert fresh["asnname"] == "Яндекс Облако"


def test_batch_cell_edit_does_not_touch_dictionary():
    """Правка ячейки у строки БЕЗ asn, с битым asn или с ASN не из
    справочника НЕ трогает справочник (записи не создаются и не меняются)."""
    a = _auth()
    pid, lid = _mk_list(a)
    client.post(f"/api/subnets/providers/{pid}/lists/{lid}/rows", headers=a,
                json={"rows": [
                    {"subnet": "203.0.113.0/24"},                    # без asn
                    {"subnet": "198.51.100.0/24", "asn": "AS999"},   # не в справочнике
                    {"subnet": "192.0.2.0/24", "asn": "мусор"},      # битый asn
                ]})
    client.post("/api/subnets/asns", headers=a,
                json={"asn": "12345", "name": "Яндекс"})
    rows = _rows(a)
    rid2 = next(x["id"] for x in rows if x["values"]["subnet"] == "198.51.100.0/24")
    r = client.patch(f"/api/subnets/providers/{pid}/lists/{lid}/rows/batch",
                     headers=a, json={"updates": [
                         {"row_id": rows[0]["id"], "col": "asnname", "value": "Новое"},
                         {"row_id": rid2, "col": "asnname", "value": "Новое"},
                         {"row_id": rid2, "col": "country", "value": "RU"},
                         {"row_id": rows[2]["id"], "col": "asn_type", "value": "hosting"},
                     ]})
    assert r.status_code == 200, r.text
    assert r.json()["updated"] == 4
    asns = client.get("/api/subnets/asns", headers=a).json()["asns"]
    assert len(asns) == 1  # только AS12345, новых записей нет
    assert asns[0]["name"] == "Яндекс"  # и не изменена
    assert "country" not in asns[0]
